import csv
import email
from io import BytesIO
import mimetypes
import os
from pyexpat.errors import messages
from urllib import request
from django.conf import settings
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from openpyxl import load_workbook
import openpyxl
from Index.forms import *
from db.models import *
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import user_passes_test,login_required
from django.core.paginator import Paginator
from django.contrib import messages 
from django.shortcuts import render
from django.db.models import Q
from openpyxl.styles import *
from openpyxl.styles.borders import Border, Side
import smtplib
from email.message import EmailMessage
from django.db.models.functions import Cast
from django.db import transaction
from django.core.mail import EmailMessage, SafeMIMEText, get_connection
from icalendar import Calendar, Event
from datetime import date, datetime, time, timedelta
from django.db.models import FloatField
import re
#LOS STATUS DEBEN SER 1="Nuevo" y 2="Completado" por temas de compatiblidad
def is_admin(user):
    return user.roles == 'Administrador'

def is_active_user(user):
    return user.is_active == 'True'

@login_required(login_url='/login/')    
def index(request):
    estados = Estado.objects.all()
    context = {
   
        'estados':estados,
    }
    return render(request, 'Index/index.html',context)

@login_required(login_url='/login/')    
def expedientesLayout(request):
    expedientes = Expediente.objects.all().order_by('-id').filter(eliminado = False)
    estatus = Estado.objects.all()
    usuarios = User.objects.exclude(roles__in=['Administrador', 'Credito', 'Gerente de Credito'])   
    usuariosCredito = User.objects.filter(roles__in=['Credito','Gerente de Credito']) 
 
    contexto = {
            'expedientes': expedientes,
            'estatus': estatus,
            'usuarios': usuarios,
            'usuariosCredito': usuariosCredito,
        }
    return render(request, 'Index/expedientesLayout.html',contexto)


def filtrar_expedientes_ajax(request):
    estatus_id = request.GET.get('estatus', '0')
    usuario_id = request.GET.get('usuarios', '0')
    usaurioCredito_id = request.GET.get('usuariosCredito', '0')
    socio_query = request.GET.get('socio', '').strip()

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    page_number = request.GET.get('page', 1)

    expedientes = Expediente.objects.all()

    if estatus_id != '0':
        expedientes = expedientes.filter(estatus_id=estatus_id)

    if usuario_id != '0':
        expedientes = expedientes.filter(usuario_id=usuario_id)

    if usaurioCredito_id != '0':
        expedientes = expedientes.filter(usuario_id=usaurioCredito_id)

    if socio_query:
        expedientes = expedientes.filter(
            Q(socio__nombre__icontains=socio_query) |
            Q(socio__id__icontains=socio_query)
        ).distinct()

    if fecha_inicio and fecha_fin:
        expedientes = expedientes.filter(fechaCreacion__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        expedientes = expedientes.filter(fechaCreacion__gte=fecha_inicio)
    elif fecha_fin:
        expedientes = expedientes.filter(fechaCreacion__lte=fecha_fin)

    expedientes = expedientes.order_by('-id').filter(eliminado = False)

    paginator = Paginator(expedientes, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        'expedientes': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator
    }

    return render(request, 'Index/tablaExpedientex.html', context)

def agregarRepresentantes(request, id):
    if request.method == 'POST':
        try:
            expediente = Expediente.objects.get(id=id)
            reps_raw = request.POST.get('representantes', '').strip().strip("|")
            nombres_reps = [x.strip() for x in reps_raw.split("||") if x.strip()]
            
            tipo_persona_map = {'F': 'Fisicas', 'M': 'Morales'}
            area_socio = tipo_persona_map.get(expediente.socio.tipoPersona)

            for nombre in nombres_reps:
                representante, created_rep = RepresentanteLegal.objects.get_or_create(
                    nombre=nombre
                )
                representante.expedientes.add(expediente)
                
                nueva, created_sec = SeccionesExpediente.objects.get_or_create(
                    expediente=expediente,
                    tipoDeSeccion='B',
                    tituloSeccion=f"Representante legal - {nombre}"
                )
                
                if created_sec:
                    _generar_con_debug_extremo(nueva, area_socio)
                
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            print(f"Error en agregarRepresentantes: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            
    return JsonResponse({'status': 'error'}, status=400)


def agregarObligados(request, id):
    if request.method == 'POST':
        try:
            expediente = Expediente.objects.get(id=id)
            obls_raw = request.POST.get('obligados', '').strip().strip("|")
            elementos_obls = [x.strip() for x in obls_raw.split("||") if x.strip()]
            
            tipo_persona_map = {'F': 'Fisicas', 'M': 'Morales'}
            area_socio = tipo_persona_map.get(expediente.socio.tipoPersona)

            for item in elementos_obls:
                try:
                    data = json.loads(item)
                    id_existente = data.get('id')
                    nombre = data.get('nombre')
                    tipo_persona = data.get('tipo_persona')
                    id_rep = data.get('representante')
                    nombre_rep_input = data.get('representante_texto')
                except json.JSONDecodeError:
                    id_existente = ""
                    nombre = item
                    tipo_persona = "F"
                    id_rep = ""
                    nombre_rep_input = ""

                if id_existente:
                    obligado = ObligadoSolidario.objects.get(id=id_existente)
                    obligado.expedientes.add(expediente)
                    tipo_persona = obligado.tipoPersona if hasattr(obligado, 'tipoPersona') else tipo_persona
                else:
                    obligado, created_obl = ObligadoSolidario.objects.get_or_create(
                        nombre=nombre,
                        defaults={'tipoPersona': tipo_persona}
                    )
                    obligado.expedientes.add(expediente)

                nueva, created_sec = SeccionesExpediente.objects.get_or_create(
                    expediente=expediente,
                    tipoDeSeccion='C',
                    tituloSeccion=f"Obligado solidario y garantes - {nombre}"
                )
                
                if created_sec:
                    _generar_con_debug_extremo(nueva, area_socio)

                if tipo_persona == 'M' and nombre_rep_input:
                    if id_rep:
                        representante = RepresentanteLegal.objects.get(id=id_rep)
                    else:
                        representante, created_rep = RepresentanteLegal.objects.get_or_create(
                            nombre=nombre_rep_input
                        )
                    
                    representante.expedientes.add(expediente)
                    obligado.representante = representante
                    obligado.save()

                    nueva_sec_rep, created_sec_rep = SeccionesExpediente.objects.get_or_create(
                        expediente=expediente,
                        tipoDeSeccion='B',
                        tituloSeccion=f"Representante legal - {nombre_rep_input}"
                    )
                    
                    if created_sec_rep:
                        _generar_con_debug_extremo(nueva_sec_rep, area_socio)
                
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            print(f"Error en agregarObligados: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error'}, status=400)

    
@login_required(login_url='/login/')
def editarExpediente(request, id):
    expediente = get_object_or_404(Expediente, pk=id)
    lineasLista = Linea.objects.filter(expediente=expediente)
    secciones = SeccionesExpediente.objects.filter(expediente=expediente).order_by('tipoDeSeccion', 'pk')
    estados = Estado.objects.all()
    usuarios = User.objects.filter(roles__in=['Ejecutivo de Servicios', 'Gerente Centro de Negocios'])
    usuariosCreditos = User.objects.filter(roles__in=['Credito','Gerente de Credito'])
    usuariosNegocios = User.objects.filter(roles__in=['Ejecutivo de Negocios'])
    rep_form = CrearRepresentante()
    obl_form = CrearObligado()
    lin_form = LineaCrearForm(expedienteInstance=expediente)
    cita_form = CitaForm(user=request.user)
    ahora_local = timezone.localtime(timezone.now())
    hoy = ahora_local.date()
    hora_actual = ahora_local.time()

    citasDisponibles = Cita.objects.filter(
        Q(dia__gt=hoy) |
        Q(dia=hoy, hora__gte=hora_actual)
    ).order_by('dia', 'hora')    

    citaAsignada = Cita.objects.filter(
        Q(expedientes=expediente),
    ).order_by('dia', 'hora').last()    
   
    if request.method == "POST":
        post = request.POST
        datos_agrupados = {}

        for key, value in post.items():
            if not key.startswith("registro-"):
                continue
           
            parts = key.split('-', 4)
            if len(parts) != 5:
                continue
           
            _, seccion_id_s, apartado_id_s, secuencial_s, field = parts
           
            try:
                ids_tuple = (int(seccion_id_s), int(apartado_id_s), int(secuencial_s))
            except ValueError:
                continue

            if ids_tuple not in datos_agrupados:
                datos_agrupados[ids_tuple] = {}

            datos_agrupados[ids_tuple][field] = value.strip()

        for (seccion_id, apartado_id, secuencial), campos in datos_agrupados.items():
            registro = RegistroSeccion.objects.filter(
                seccion__id=seccion_id,
                seccion__expediente=expediente,
                apartado__id=apartado_id,
                secuencial=secuencial
            ).first()
           
            if not registro:
                continue
           
            if registro.seccion.tipoDeSeccion == 'VII' and 'descripcion_libre' in campos:
                registro.comentario = campos['descripcion_libre']
            elif 'comentario' in campos:
                registro.comentario = campos['comentario']

            if 'estatus' in campos:
                registro.estatus = campos['estatus']

            if 'comentarioCredito' in campos:
                registro.comentarioCredito = campos['comentarioCredito']
           
            if 'es_fecha' in campos:
                registro.es_fecha = campos['es_fecha'].lower() == 'true'

            if registro.es_fecha:
                registro.numero = None
                if 'fecha_date' in campos:
                    val_fecha = campos['fecha_date']
                    if val_fecha:
                        try:
                            registro.fecha = datetime.strptime(val_fecha, "%Y-%m-%d").date()
                        except ValueError:
                            registro.fecha = None
                    else:
                        registro.fecha = None
                else:
                    registro.fecha = None
            else:
                registro.fecha = None
                if 'fecha_num' in campos:
                    val_numero = campos['fecha_num']
                    if val_numero:
                        try:
                            registro.numero = val_numero
                        except ValueError:
                            registro.numero = None
                    else:
                        registro.numero = None
           
            registro.save()

        messages.success(request, 'Datos guardados con éxito.')
        return redirect(reverse('Index:editarExpediente', args=[expediente.pk]))

    num_kepler = expediente.socio.numeroKepler
    if num_kepler is None or str(num_kepler).strip() == "" or str(num_kepler).strip() == "0000" or str(num_kepler).strip() == "0":
        stringSocio = f"0000 {expediente.socio.nombre}"
    else:
        stringSocio = str(num_kepler).strip()

    archivos_encontrados = checarRuta(stringSocio, secciones)    
    print(stringSocio)
    context = {
        'estados': estados,
        'usuarios': usuarios,
        'expediente': expediente,
        'lineasLista': lineasLista,
        'secciones': [],
        'citasDisponibles': citasDisponibles,
        'citaAsignada': citaAsignada,
        'usuariosCreditos': usuariosCreditos,
        'usuariosNegocios': usuariosNegocios,
    }
   
    totalRegistros = 0
    totalRegistrosLlenos = 0

    es_credito = request.user.roles in ['Credito', 'Gerente de Credito']
    estatus_recepcion = expediente.estatus.nombre == "Completo"

    for seccion in secciones:
        registros_existentes = RegistroSeccion.objects.filter(seccion=seccion).select_related('apartado').order_by('apartado__clave', 'secuencial')
       
        filas = []
        for registro in registros_existentes:
            if es_credito and estatus_recepcion:
                if not registro.estatus or registro.estatus.strip() == "" or registro.estatus.strip().upper() == "N/A":
                    continue

            totalRegistros += 1

            if registro.estatus and registro.estatus != "":
                totalRegistrosLlenos += 1

            fecha_html = ""
            if registro.es_fecha:
                if registro.fecha:
                    fecha_html = registro.fecha.strftime("%Y-%m-%d")
            else:
                if registro.numero is not None:
                    fecha_html = str(registro.numero)
           
            info_archivo = archivos_encontrados.get(registro.id)
           
            filas.append({
                'apartado': registro.apartado,
                'registro': registro,
                'fecha_html': fecha_html,
                'archivo_url': info_archivo['ruta'] if info_archivo else None
            })
           
        context['secciones'].append({
            'seccion': seccion,
            'filas': filas,
        })
       
    context['totalRegistros'] = totalRegistros
    context['totalRegistrosLlenos'] = totalRegistrosLlenos
    context['rep_form'] = rep_form
    context["obl_form"] = obl_form
    context["lin_form"] = lin_form
    context['cita_form'] = cita_form
    lista_reps = RepresentanteLegal.objects.filter(expedientes=expediente)
    lista_obls = ObligadoSolidario.objects.filter(expedientes=expediente)
    context['lista_reps'] = lista_reps
    context["lista_obls"] = lista_obls
    return render(request, 'Index/editarExpediente.html', context)


def checarRuta(identificador_socio, secciones):
    #print("\n================= DEBUG CHECAR RUTA =================")
    rutaServidor = fr"\\192.168.0.96\intranetucg$$\Evidencias\652 Digitalización de expedientes de crédito"
    #print(f"[*] Buscando identificador_socio: '{identificador_socio}'")

    carpeta_socio = None
    try:
        dirs_servidor = os.listdir(rutaServidor)
        for nombre_dir in dirs_servidor:
            if identificador_socio in nombre_dir:
                carpeta_socio = os.path.join(rutaServidor, nombre_dir)
                break
    except Exception as e:
        #print(f"[-] ERROR al listar la raíz del servidor: {e}")
        return {}

    if not carpeta_socio:
        return {}

    rutaMaestra = os.path.join(carpeta_socio, "Maestra")
    rutaOperativa = os.path.join(carpeta_socio, "Operativa")

    mapeo_secciones = {
        "1": os.path.join(rutaMaestra, "I. Identificación del Socio"),
        "2": os.path.join(rutaMaestra, "II. Información Financiera"),
        "3": os.path.join(rutaOperativa, "III. Estudio de Crédito"),
        "4": os.path.join(rutaOperativa, "IV. Información de garantias"),
        "5": os.path.join(rutaOperativa, "V. Contratos"),
        "6": os.path.join(rutaOperativa, "VI. Seguimiento"),
        "7": os.path.join(rutaOperativa, "VII. Correspondencia")
    }
    
    meses_es = {
        1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
        7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic"
    }
    
    resultados = {}

    for seccion in secciones:
        registros_existentes = RegistroSeccion.objects.filter(seccion=seccion).select_related('apartado', 'seccion__expediente')
        
        for registro in registros_existentes:
            clave = str(registro.apartado.clave).strip()
            prefijo = clave.split('.')[0]
            
            ruta_busqueda = mapeo_secciones.get(prefijo)
            #print(f"\n--- Procesando Registro ID: {registro.id} ---")
            #print(f"    Clave DB: {clave} | Prefijo: {prefijo}")

            if not ruta_busqueda:
                continue
                
            if prefijo in ["3", "4", "5", "6", "7"] and not os.path.exists(ruta_busqueda):
                try:
                    expediente_actual = registro.seccion.expediente
                    lineas_asociadas = Linea.objects.filter(expediente=expediente_actual)
                    
                    if os.path.exists(rutaOperativa):
                        subdirs_operativa = os.listdir(rutaOperativa)
                        enrutado_exitoso = False
                        
                        for linea in lineas_asociadas:
                            prefijo_linea = f"{linea.numero} "
                            #print(f"    [*] Buscando carpeta de línea que inicie con: '{prefijo_linea}'")
                            
                            for subdir in subdirs_operativa:
                                if subdir.strip().startswith(prefijo_linea):
                                    ruta_subdir_completa = os.path.join(rutaOperativa, subdir)
                                    if os.path.isdir(ruta_subdir_completa):
                                        nombre_carpeta_seccion = os.path.basename(mapeo_secciones.get(prefijo))
                                        posible_ruta = os.path.join(ruta_subdir_completa, nombre_carpeta_seccion)
                                        if os.path.exists(posible_ruta):
                                            ruta_busqueda = posible_ruta
                                            #print(f"    [+] Enrutamiento correcto hacia: {ruta_busqueda}")
                                            enrutado_exitoso = True
                                            break
                            if enrutado_exitoso:
                                break
                except Exception as ex_op:
                    print(f"    [-] Error escaneando Operativa: {ex_op}")

            if not os.path.exists(ruta_busqueda):
                #print(f"    [-] Cancelado: La ruta final NO existe: {ruta_busqueda}")
                continue
                
            archivo_mas_reciente = None
            mtime_maximo = 0

            opciones_clave = [clave]
            partes_clave = clave.split('.')
            if len(partes_clave) == 2:
                try:
                    num_principal = partes_clave[0]
                    num_decimal = int(partes_clave[1])
                    opciones_clave.append(f"{num_principal}.{num_decimal:02d}")
                    opciones_clave.append(f"{num_principal}.{num_decimal}")
                except ValueError:
                    pass
            
            opciones_clave = list(set(opciones_clave))
            #print(f"    [*] Variantes de clave a buscar en disco: {opciones_clave}")

            try:
                items_directorio = os.listdir(ruta_busqueda)
                for nombre_item in items_directorio:
                    ruta_item = os.path.join(ruta_busqueda, nombre_item)
                    
                    if os.path.isdir(ruta_item):
                        nombre_item_clean = nombre_item.strip()
                        
                        condicion_carpeta = any(
                            nombre_item_clean.startswith(opt + " ") or 
                            nombre_item_clean.startswith(opt + ".") or
                            nombre_item_clean == opt
                            for opt in opciones_clave
                        )
                        
                        if condicion_carpeta:
                            archivos_internos = os.listdir(ruta_item)
                       
                            for archivo in archivos_internos:
                                cumple_inicio_archivo = any(archivo.startswith(opt) for opt in opciones_clave)
                                if not cumple_inicio_archivo:
                                    continue
                                
                                archivo_lower = archivo.lower()
                                coincide = False
                                
                                if registro.es_fecha:
                                    if registro.fecha:
                                        mes_texto = meses_es.get(registro.fecha.month, "")
                                        anio_dos_digitos = registro.fecha.strftime("%y")
                                        if mes_texto in archivo_lower and anio_dos_digitos in archivo_lower:
                                            coincide = True
                                else:
                                    if registro.numero:
                                        criterio_num = str(registro.numero).strip().lower()
                                        if criterio_num in archivo_lower:
                                            coincide = True
                                
                                if coincide:
                                    ruta_archivo = os.path.join(ruta_item, archivo)
                                    if os.path.isfile(ruta_archivo):
                                        mtime = os.path.getmtime(ruta_archivo)
                                        if mtime > mtime_maximo:
                                            mtime_maximo = mtime
                                            archivo_mas_reciente = ruta_archivo
            except Exception:
                continue

            if archivo_mas_reciente:
                #print(f"    [=>] Asignado archivo ganador: {archivo_mas_reciente}")
                resultados[registro.id] = {
                    'clave': clave,
                    'ruta': archivo_mas_reciente,
                    'fecha_modificacion': datetime.fromtimestamp(mtime_maximo)
                }

    #print("\n================= FIN DEBUG CHECAR RUTA =================\n")
    return resultados

@login_required(login_url='/login/')
def lineaCrear(request, id):
    expedienteInstance = get_object_or_404(Expediente, pk=id)
    
    if request.method == "POST":
        lineasData = request.POST.get('lineas', '')
        
        if lineasData:
            listaSegmentada = lineasData.split('||')
            tipo_persona_map = {'F': 'Fisicas', 'M': 'Morales'}
            area_socio = tipo_persona_map.get(expedienteInstance.socio.tipoPersona)
            
            secciones_restantes = ['III', 'IV', 'V', 'VI','VII']
            
            try:
                with transaction.atomic():
                    for bloque in listaSegmentada:
                        if not bloque.strip():
                            continue
                        
                        partes = bloque.split('::')
                        if len(partes) < 4:
                            continue
                            
                        numero = partes[0].strip()
                        monto = partes[1].strip()
                        tipoLineaId = partes[2].strip()
                        vigente = partes[3] == 'true'
                        
                        nueva_linea = Linea.objects.create(
                            expediente=expedienteInstance,
                            numero=numero,
                            monto=monto,
                            tipoLinea=tipoLineaId,
                            vigente=True
                        )
                        
                        for tipo_sec in secciones_restantes:
                            nueva_seccion = SeccionesExpediente.objects.create(
                                expediente=expedienteInstance,
                                linea=nueva_linea,
                                tipoDeSeccion=tipo_sec
                            )
                            
                            _generar_con_debug_extremo(nueva_seccion, area_socio)
                            
                return JsonResponse({'status': 'success'}, status=200)
                
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            
        return JsonResponse({'status': 'error', 'message': 'No data'}, status=400)
        
    return JsonResponse({'status': 'error'}, status=405)

@login_required(login_url='/login/')

def servirArchivo(request):
    ruta_archivo = request.GET.get('ruta')
    if not ruta_archivo or not os.path.exists(ruta_archivo):
        raise Http404("El archivo no existe")

    nombre_archivo = os.path.basename(ruta_archivo)
    mime_type, _ = mimetypes.guess_type(ruta_archivo)
    
    response = FileResponse(open(ruta_archivo, 'rb'), content_type=mime_type)
    response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'
    return response


        
        
@login_required(login_url='/login/')
def expediente_cambiar_usuario(request, id):
    expediente = get_object_or_404(Expediente, pk=id)

    if request.method == "POST":
        nuevo_usuarios_id = request.POST.get("usuario")

        if nuevo_usuarios_id:
            expediente.usuario_id = nuevo_usuarios_id
            expediente.save()


    return redirect('Index:editarExpediente', id=expediente.id)
        
@login_required(login_url='/login/')
def cambiarUsuarioCredito(request, id):
    expediente = get_object_or_404(Expediente, pk=id)

    if request.method == "POST":
        nuevo_usuarios_id = request.POST.get("usuario")

        if nuevo_usuarios_id:
            expediente.usuarioCredito_id = nuevo_usuarios_id
            expediente.save()


    return redirect('Index:editarExpediente', id=expediente.id)
@login_required(login_url='/login/')
def expediente_llenar(request, id):
    expediente = get_object_or_404(Expediente, pk=id)
    secciones = SeccionesExpediente.objects.filter(expediente=expediente).order_by('tipoDeSeccion', 'pk')
    for seccion in secciones:
        apartados = ApartadoCatalogo.objects.filter(tipoDeSeccion=seccion.tipoDeSeccion).order_by('clave')
        for apartado in apartados:
            registro = RegistroSeccion.objects.filter(seccion=seccion, apartado=apartado).first()
            if registro:
                if not registro.estatus:
                    registro.estatus = 'N/A'
                else:
                    registro.estatus = registro.estatus
                registro.save()
    return redirect('Index:editarExpediente', id=expediente.id)


@login_required(login_url='/login/')
def crearExpediente(request):
    if request.method == "POST":
        exp_form = ExpedienteCrearForm(request.POST)
        rep_form = CrearRepresentante(request.POST)
        obl_form = CrearObligado(request.POST)

        if exp_form.is_valid() and rep_form.is_valid() and obl_form.is_valid():
            try:
                with transaction.atomic():
                    socio_existente = exp_form.cleaned_data.get('socio')
                    nombre_manual = exp_form.cleaned_data.get('socio_manual_nombre')
                    tipo_manual = exp_form.cleaned_data.get('socio_manual_tipo')
                    numeroK = exp_form.cleaned_data.get('socio_manual_numero')

                    socio_a_asignar = None
                    if socio_existente:
                        socio_a_asignar = socio_existente
                    elif nombre_manual and tipo_manual:
                        socio_a_asignar = Socio.objects.create(
                            nombre=nombre_manual, 
                            tipoPersona=tipo_manual, 
                            numeroKepler=numeroK
                        )
                    
                    if not socio_a_asignar:
                        raise ValueError("Debe seleccionar un socio existente o completar los datos manuales.")

                    tipo_persona = socio_a_asignar.tipoPersona
                    tipo_persona_map = {'F': 'Fisicas', 'M': 'Morales'}
                    area_socio = tipo_persona_map.get(tipo_persona)

                    reps_raw = rep_form.cleaned_data.get('representantes', '').strip().strip("|")
                    nombres_reps = [x.strip() for x in reps_raw.split("||") if x.strip()]
                    
                    obls_raw = obl_form.cleaned_data.get('obligados', '').strip().strip("|")
                    raw_items = [x.strip() for x in obls_raw.split("||") if x.strip()]

                    import json
                    listado_obligados_procesados = []

                    for item in raw_items:
                        try:
                            data_obl = json.loads(item)
                            listado_obligados_procesados.append(data_obl)
                        except json.JSONDecodeError:
                            listado_obligados_procesados.append({
                                "id": "",
                                "nombre": item,
                                "tipo_persona": "F",
                                "tipo_persona_texto": "Persona Física",
                                "representante": "",
                                "representante_texto": ""
                            })

                    if tipo_persona == 'M' and not nombres_reps:
                        raise ValueError("Las Personas Morales requieren representantes.")

                    expediente = exp_form.save(commit=False)
                    expediente.socio = socio_a_asignar
                    expediente.estatus_id = 1
                    #expediente.usuario = request.user
                    expediente.save()
                    
                    EstadosFechas.objects.create(
                        expediente=expediente,
                        estado=Estado.objects.get(nombre="Nuevo"),
                        fecha=datetime.now(),
                        usuario=request.user
                    )
                    
                    for rep_nombre in nombres_reps:
                        representante, created = RepresentanteLegal.objects.get_or_create(nombre=rep_nombre)
                        representante.expedientes.add(expediente)

                    for o_data in listado_obligados_procesados:
                        if o_data.get("id"):
                            obligado = ObligadoSolidario.objects.filter(id=o_data["id"]).first()
                        else:
                            obligado = ObligadoSolidario.objects.filter(nombre=o_data["nombre"]).first()
                            if not obligado:
                                t_persona = o_data.get("tipo_persona") if o_data.get("tipo_persona") else "F"
                                obligado = ObligadoSolidario.objects.create(
                                    nombre=o_data["nombre"], 
                                    tipoPersona=t_persona
                                )
                        
                        if obligado:
                            obligado.expedientes.add(expediente)
                            
                            if obligado.tipoPersona == 'M' and o_data.get("representante_texto"):
                                rep_nom = o_data["representante_texto"]
                                r_legal, _ = RepresentanteLegal.objects.get_or_create(nombre=rep_nom)
                                obligado.representante = r_legal
                                integrado = r_legal
                                integrado.expedientes.add(expediente)
                                obligado.save()

                    secciones_permitidas = ['A', 'B', 'C', 'I', 'II']
                    SECCIONES = SeccionesExpediente.SECCIONES
                    
                    for tipo_sec, titulo_sec in SECCIONES:
                        if tipo_sec not in secciones_permitidas:
                            continue
                            
                        if tipo_sec == 'B':
                            for nombre in nombres_reps:
                                nueva = SeccionesExpediente.objects.create(
                                    expediente=expediente, 
                                    tipoDeSeccion=tipo_sec, 
                                    tituloSeccion=f"{titulo_sec} - {nombre}"
                                )
                                _generar_con_debug_extremo(nueva, area_socio)
                        elif tipo_sec == 'C':
                            for o_data in listado_obligados_procesados:
                                nombre_obl = o_data["nombre"]
                                nueva = SeccionesExpediente.objects.create(
                                    expediente=expediente, 
                                    tipoDeSeccion=tipo_sec, 
                                    tituloSeccion=f"{titulo_sec} - {nombre_obl}"
                                )
                                _generar_con_debug_extremo(nueva, area_socio)
                                
                                if o_data.get("id"):
                                    obl_obj = ObligadoSolidario.objects.filter(id=o_data["id"]).first()
                                else:
                                    obl_obj = ObligadoSolidario.objects.filter(nombre=nombre_obl).first()
                                    
                                if obl_obj and obl_obj.tipoPersona == 'M' and obl_obj.representante:
                                    nueva_rep = SeccionesExpediente.objects.create(
                                        expediente=expediente,
                                        tipoDeSeccion='B',
                                        tituloSeccion=f"Rep. Legal Obligado - {obl_obj.representante.nombre}"
                                    )
                                    _generar_con_debug_extremo(nueva_rep, area_socio)
                        else:
                            nueva = SeccionesExpediente.objects.create(
                                    expediente=expediente, 
                                    tipoDeSeccion=tipo_sec
                                )
                            _generar_con_debug_extremo(nueva, area_socio)

                return JsonResponse({
                    'success': True, 
                    'redirect_url': reverse('Index:expedientesLayout')
                })

            except ValueError as e:
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
            except Exception as e:
                return JsonResponse({
                    'success': False, 
                    'error': f"Ocurrió un error inesperado: {str(e)}"
                }, status=500)
        else:
            error_msg = f"Errores: {exp_form.errors.as_text()} {rep_form.errors.as_text()} {obl_form.errors.as_text()}"
            return JsonResponse({'success': False, 'error': error_msg}, status=400)

    exp_form = ExpedienteCrearForm()
    rep_form = CrearRepresentante()
    obl_form = CrearObligado()
    return render(request, 'Index/crearExpediente.html', {
        'exp_form': exp_form, 
        'rep_form': rep_form, 
        'obl_form': obl_form
    })

def _generar_con_debug_extremo(seccion_obj, area_socio, linea_instance=None):
    filtro_permitido = [area_socio, 'Ambas']
    
    todos_los_apartados = ApartadoCatalogo.objects.filter(
        tipoDeSeccion=seccion_obj.tipoDeSeccion
    ).annotate(
        clave_numerica=Cast('clave', FloatField())
    ).order_by('clave_numerica')

    for ap in todos_los_apartados:
        area_en_db = str(ap.areaDondeAplica).strip()
        condicion = area_en_db in filtro_permitido
        
        if condicion:
            RegistroSeccion.objects.create(
                seccion=seccion_obj, 
                apartado=ap
            )
@login_required(login_url='/login/')    
def expediente_eliminar(request,id):
    expediente = Expediente.objects.get(id=id)
    #expediente.delete()
    expediente.eliminado = True
    expediente.save()
    return redirect('Index:expedientesLayout')



def correoParaRevision(expediente,reenviado):
    destinatario = ["portiz@ucg.com.mx","dCorrea@ucg.com.mx", "mrubio@ucg.com.mx",]#"daudiffred@ucg.com.mx"
    dominio = "http://192.168.0.29:8000/expedientes/editarExpediente/"
    urlFinal = f"{dominio}{expediente.id}/"
    
    usuarioNombre = expediente.usuario.username 
    if reenviado =='True':

        asunto = "Expediente corregido para revisión"
        cuerpoHtml = f"""
    <html>
        <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; padding: 20px; border-radius: 10px;">
                <h2 style="color: #2c3e50;">Actualización de estatus del expediente</h2>
                <p>Estimado usuario,</p>
                <p>Le informamos que el expediente <strong>{expediente.id}</strong> del socio <strong>{expediente.socio.nombre}</strong> ha sido corregido y está listo para su revisión.</p>
                <div style="margin: 30px 0; text-align: center;">
                    <a href="{urlFinal}" 
                       style="background-color: #007bff; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                        Revisar Expediente
                    </a>
                </div>
                <p style="font-size: 0.9em; color: #555;">
                    Atentamente,<br>
                    <strong>{usuarioNombre}</strong>
                </p>
                <hr style="border: 0; border-top: 1px solid #eee;">
                <p style="font-size: 0.8em; color: #999; text-align: center;">
                    Este es un mensaje automático, por favor no responda a este correo.
                </p>
            </div>
        </body>
    </html>
    """
    else:    
        asunto = "Expediente listo para revisión"
        cuerpoHtml = f"""

   <html>
        <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; padding: 20px; border-radius: 10px;">
                <h2 style="color: #2c3e50;">Actualización de estatus del expediente</h2>
                <p>Estimado usuario,</p>
                <p>Le informamos que el expediente <strong>{expediente.id}</strong> del socio <strong>{expediente.socio.nombre}</strong> está listo para su revisión.</p>
                <div style="margin: 30px 0; text-align: center;">
                    <a href="{urlFinal}" 
                       style="background-color: #007bff; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                        Revisar Expediente
                    </a>
                </div>
                <p style="font-size: 0.9em; color: #555;">
                    Atentamente,<br>
                    <strong>{usuarioNombre}</strong>
                </p>
                <hr style="border: 0; border-top: 1px solid #eee;">
                <p style="font-size: 0.8em; color: #999; text-align: center;">
                    Este es un mensaje automático, por favor no responda a este correo.
                </p>
            </div>
        </body>
    </html>
    """
    SMTP_HOST = "ucg.com.mx"
    SMTP_PORT = 587
    USUARIO = "informacion@ucg.com.mx"
    CONTRASENA = "UcG911_@!#"

    mensaje = email.message.EmailMessage()
    mensaje["From"] = USUARIO
    mensaje["To"] = ", ".join(destinatario) if isinstance(destinatario, list) else destinatario
    mensaje["Subject"] = asunto
    
    mensaje.add_alternative(cuerpoHtml, subtype="html")

    try:
        servidor = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        servidor.starttls()
        servidor.login(USUARIO, CONTRASENA)
        servidor.send_message(mensaje)
        servidor.quit()
        print("Correo enviado correctamente")
    except Exception as e:
        print("Error al enviar correo:", e)
    #return redirect('Index:editarExpediente', expediente.id)

@login_required(login_url='/login/')    
def rechazarExpediente(request,expedienteID):
    from email.message import EmailMessage
    expediente = get_object_or_404(Expediente, pk=expedienteID)
    getEstado = Estado.objects.get(nombre='Rechazado')
    expediente.estatus = getEstado
    expediente.save()
    darAlta(expediente,getEstado.nombre,request.user)

    seccionesConComentarios = []
    secciones = SeccionesExpediente.objects.filter(expediente=expediente).order_by('tipoDeSeccion', 'pk')
    
    for seccion in secciones:
        apartados = ApartadoCatalogo.objects.filter(tipoDeSeccion=seccion.tipoDeSeccion).order_by('clave')
        for apartado in apartados:
            registro = RegistroSeccion.objects.filter(seccion=seccion, apartado=apartado).first()
            
            if registro is not None:
                if registro.comentarioCredito:
                    seccionesConComentarios.append({
                        'nombreSeccion': str(seccion.tituloSeccion),
                        'claveApartado': str(apartado.clave),
                        'comentario': str(registro.comentarioCredito)
                    })
                    
    destinatario = [expediente.usuario.email] #expediente.usuarioCredito.email
    #destinatario = ["daudiffred@ucg.com.mx"]    
    dominio = "http://192.168.0.29:8000/expedientes/editarExpediente/"
    url_final = f"{dominio}{expediente.id}/"
    asunto = f"RECHAZO: Expediente {expediente.id} - {expediente.socio.nombre}"
    
    comentarios_texto = ""
    for item in seccionesConComentarios:
        comentarios_texto += f"\n- [{item['nombreSeccion']} - {item['claveApartado']}]: {item['comentario']}"


    cuerpo_html = f"""
    <html>
        <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; padding: 20px; border-radius: 10px;">
                <h2 style="color: #2c3e50;">Actualización de estatus del expediente</h2>
                <p>Estimado usuario,</p>
                <p>Le informamos que el expediente {expediente.id} del socio {expediente.socio.nombre} - {expediente.socio.numeroKepler} ha sido rechazado.</p>
                <p>Atte {expediente.usuarioCredito.username}.</p>
                <div style="margin: 30px 0; text-align: center;">
                    <a href="{url_final}" 
                       style="background-color: #007bff; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                        Revisar Expediente
                    </a>
                </div>
                <div style="background-color: #fdf2f2; border-left: 4px solid #f5c6cb; padding: 15px; margin: 20px 0; border-radius: 4px;">
                    <h4 style="margin-top: 0; color: #a94442;">Detalle de observaciones por sección:</h4>
                    <ul style="margin-bottom: 0; padding-left: 20px;">
    """

    for item in seccionesConComentarios:
        cuerpo_html += f"""
                        <li style="margin-bottom: 10px;">
                            <strong>{item['nombreSeccion']} (Clave: {item['claveApartado']}):</strong> {item['comentario']}
                        </li>
        """

    cuerpo_html += f"""
                    </ul>
                </div>
          
                <hr style="border: 0; border-top: 1px solid #eee;">
                <p style="font-size: 0.8em; color: #999; text-align: center;">
                    Este es un mensaje automático, por favor no responda a este correo.
                </p>
            </div>
        </body>
    </html>
    """
    SMTP_HOST = "ucg.com.mx"
    SMTP_PORT = 587
    USUARIO = "informacion@ucg.com.mx"
    CONTRASENA = "UcG911_@!#"
    
    mensaje = EmailMessage()
    mensaje["From"] = USUARIO
    mensaje["To"] = ", ".join(destinatario) if isinstance(destinatario, list) else destinatario
    mensaje["Subject"] = asunto
    
    mensaje.add_alternative(cuerpo_html, subtype="html")

    try:
        servidor = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        servidor.starttls()
        servidor.login(USUARIO, CONTRASENA)
        servidor.send_message(mensaje)
        servidor.quit()
        print("Correo enviado correctamente")
    except Exception as e:
        print("Error al enviar correo:", e)
    return redirect('Index:editarExpediente', expediente.id)


def darAlta(expediente, nuevo_estatus_nombre, usuario):
    #estadoActual = EstadosFechas.objects.filter(
    #    expediente=expediente, 
    #    estado__nombre=nuevo_estatus_nombre
    #).first()
    #
    #if estadoActual:
    #    estadoActual.fecha = datetime.now()
    #    estadoActual.save()
    #else:    
    EstadosFechas.objects.create(
            expediente=expediente,
            estado=Estado.objects.get(nombre=nuevo_estatus_nombre),
            fecha=datetime.now(),
            usuario=usuario
        )
@login_required(login_url='/login/')
def cambiarEstado(request, id):
    expediente = get_object_or_404(Expediente, pk=id)

    if request.method == "POST":
        nuevo_estatus_nombre = request.POST.get("estatus")
        if nuevo_estatus_nombre and nuevo_estatus_nombre != str(expediente.estatus.nombre):
            #darAlta(expediente, nuevo_estatus_nombre, request.user)
            #if nuevo_estatus_nombre == "Completo":
            #    correoParaRevision(expediente)
            #if nuevo_estatus_nombre == 'Rechazado':
            #    mandarCorreoRechazo(expediente.id)
            
            getEstado = Estado.objects.get(nombre=nuevo_estatus_nombre)
            expediente.estatus = getEstado
            expediente.save()
            
    return redirect('Index:editarExpediente', id=expediente.id)

@login_required(login_url='/login/')    
def editar_layout(request):
    formset = EstadoFormSet(queryset=Estado.objects.all())
    formSocios = EditarSocio()
    formApartado = ModificarApartado()
    formSelectorApartado = SelectorApartadoForm()

    if request.method == 'POST':
        form_name = request.POST.get('form_name')

        if form_name == 'estado_form':
            formset = EstadoFormSet(request.POST)
            if formset.is_valid():
                formset.save()
                messages.success(request, "Los estados se actualizaron con éxito.")
                return redirect('Index:editar_layout')
            else:
                messages.error(request, f"Error en estados: {formset.errors.as_text()}")
        
        elif form_name == 'socio_form':
            socio_id = request.POST.get('socio_id')
            if socio_id and socio_id.strip() != "":
                instance = get_object_or_404(Socio, pk=socio_id)
                formSocios = EditarSocio(request.POST, instance=instance)
                accion = "actualizado"
            else:
                formSocios = EditarSocio(request.POST)
                accion = "creado"
                
            if formSocios.is_valid():
                formSocios.save()
                messages.success(request, f"Socio {accion} con éxito.")
                return redirect('Index:editar_layout')
            else:
                errores = " / ".join([f"{v[0]}" for k, v in formSocios.errors.items()])
                messages.error(request, f"Error al guardar socio: {errores}")
        
        elif form_name == 'apartado_form':
            apartado_id = request.POST.get('apartado_id')
            if apartado_id and apartado_id.strip() != "":
                instance = get_object_or_404(ApartadoCatalogo, pk=apartado_id)
                formApartado = ModificarApartado(request.POST, instance=instance)
                accion = "actualizado"
            else:
                formApartado = ModificarApartado(request.POST)
                accion = "creado"
                
            if formApartado.is_valid():
                formApartado.save()
                messages.success(request, f"Apartado de catálogo {accion} con éxito.")
                return redirect('Index:editar_layout')
            else:
                errores = " / ".join([f"{v[0]}" for k, v in formApartado.errors.items()])
                messages.error(request, f"Error en el apartado: {errores}")

    context = {
        'formset': formset,
        'formSocios': formSocios,
        'formApartado': formApartado,
        'formSelectorApartado': formSelectorApartado,
    }
    return render(request, 'Index/ajustes.html', context)

def obtener_apartado_data(request, apartado_id):
    apartado = get_object_or_404(ApartadoCatalogo, pk=apartado_id)
    data = {
        'tipoDeSeccion': apartado.tipoDeSeccion,
        'clave': apartado.clave,
        'descripcion': apartado.descripcion,
        'areaDondeAplica': apartado.areaDondeAplica,
    }
    return JsonResponse(data)


def obtener_obligado_data(request, obligado_id):
    obligado = get_object_or_404(ObligadoSolidario, id=obligado_id)
    
    rep_id = ""
    rep_nombre = ""
    
    if preparado := obligado.representante:
        rep_id = preparado.id
        rep_nombre = preparado.nombre
            
    return JsonResponse({
        'tipoPersona': obligado.tipoPersona,
        'representante_id': rep_id,
        'representante_nombre': rep_nombre
    })
def obtener_socio_data(request, socio_id):
    try:
        socio = Socio.objects.get(pk=socio_id)
        data = {
            'nombre': socio.nombre,
            'tipoPersona': socio.tipoPersona, 
            'numeroKepler': socio.numeroKepler, 

        }
        return JsonResponse(data)
    except Socio.DoesNotExist:
        return JsonResponse({'error': 'Socio no encontrado'}, status=404)

def obtener_lineas_socio(request, socio_id):
    socio = Socio.objects.get(pk=socio_id)
    socio_id = request.GET.get('socio_id')
    lineas = Linea.objects.filter(socio_id=socio).values('id', 'numero', 'monto')
    return JsonResponse(list(lineas), safe=False)

@login_required(login_url='/login/')
def exportarExcel(request, id):
    expediente = get_object_or_404(Expediente, pk=id)
    
    ruta_directorio = fr"C:\Users\Diego Audiffred\Downloads\Lista de proyectos\PlataformaExpedientes\drAlejandro\static"
    ruta_plantilla = os.path.join(ruta_directorio, "FormatoParaExpedientes.xlsx")
    
    try:
        wb = load_workbook(ruta_plantilla)
        ws = wb['Template1'] if 'Template1' in wb.sheetnames else wb.active

        ws['A2'] = f" {expediente.socio.nombre} -# {expediente.id}"

        fila_actual = 4
        
        secciones = SeccionesExpediente.objects.filter(expediente=expediente).order_by('tipoDeSeccion', 'pk')

        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment_center = Alignment(horizontal='center', wrap_text=True, vertical='center')

        for item in secciones:
            registros = RegistroSeccion.objects.filter(seccion=item).select_related('apartado').order_by('apartado__clave')
            
            if not registros.exists():
                continue

            ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=5)
            
            for col in range(1, 6):
                celda_borde = ws.cell(row=fila_actual, column=col)
                celda_borde.border = thin_border
                if col == 1:
                    celda_borde.value = str(item.tituloSeccion).upper()
                    celda_borde.alignment = alignment_center
                    celda_borde.font = Font(bold=True)
                    celda_borde.fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
                else:
                    celda_borde.fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")

            fila_actual += 1

            for reg in registros:
                datos = [
                    reg.apartado.clave,
                    reg.apartado.descripcion,
                    reg.fecha.strftime("%d/%m/%Y") if reg.fecha else "",
                    reg.estatus,
                    reg.comentario
                ]

                for col_idx, valor in enumerate(datos, start=1):
                    cell = ws.cell(row=fila_actual, column=col_idx, value=valor)
                    cell.border = thin_border
                    cell.alignment = alignment_center
                
                fila_actual += 1

        nombre_socio = "".join([c for c in str(expediente.socio.nombre) if c.isalnum() or c==' ']).replace(' ', '_')
        nombre_salida = f"Expediente_{expediente.id}_{nombre_socio}.xlsx"
        ruta_salida = os.path.join(ruta_directorio, nombre_salida)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        with open(ruta_salida, "wb") as f:
            f.write(buffer.getbuffer())

        buffer.seek(0)
        response = FileResponse(buffer, as_attachment=True, filename=nombre_salida)
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    except Exception as e:
        return HttpResponse(f"Error critico al generar archivo: {str(e)}", status=500)
    
@login_required(login_url='/login/')
def exportarPDF(request, id):
    expediente = get_object_or_404(Expediente, pk=id)
    
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    
    try:
        secciones_data = []
        secciones = SeccionesExpediente.objects.filter(expediente=expediente).order_by('tipoDeSeccion', 'pk')

        for item in secciones:
            registros = RegistroSeccion.objects.filter(seccion=item).select_related('apartado').order_by('apartado__clave')
            
            if registros.exists():
                secciones_data.append({
                    'titulo': str(item.tituloSeccion).upper(),
                    'registros': registros
                })

        context = {
            'expediente': expediente,
            'secciones': secciones_data,
            'socio_info': f"{expediente.socio.nombre} - #{expediente.id}"
        }

        ruta_logo = r"C:/Users/Diego Audiffred/Downloads/Lista de proyectos/PlataformaExpedientes/drAlejandro/static/img/Logo.png"

        html_string = f"""
        <html>
        <head>
            <style>
                @page {{ size: letter; margin: 1cm; }}
                body {{ font-family: Arial, sans-serif; font-size: 10px; }}
.header {{ 
    text-align: center; 
    font-weight: bold; 
    font-size: 24px; 
    background-color: #d6dce4; 
    border: 1px solid black; 
    padding: 15px 5px; /* El primer valor (30px) centra verticalmente */
}}                table {{ width: 100%; border-collapse: collapse; margin-bottom: 0px; table-layout: fixed; }}
                th, td {{ border: 1px solid black; padding: 4px; word-wrap: break-word; }}
                .titulo-seccion {{ background-color: #9BC2E6; font-weight: bold; text-align: center; }}
                .col-clave {{ width: 10%; text-align: center; }}
                .col-desc {{ width: 40%; }}
                .col-fecha {{ width: 12%; text-align: center;}}
                .col-status {{ width: 12%; text-align: center; }}
                .col-com {{ width: 26%; }}
            </style>
        </head>
        <body>
            <div style="background-color:#3eb1c8; border: 1px solid black; padding: 5px;">
                <table style="width: 100%; border: none; margin: 0;background-color:#3eb1c8; border: 1px solid black;">
                    <tr>
                        <td style="width: 20%; border: none; vertical-align: middle; text-align: left;">
                            <img src="{ruta_logo}" style="height: 40px;width:120px">
                        </td>
                        <td style="width: 80%; border: none; vertical-align: middle; text-align: right; font-weight: bold; font-size: 18px;color:#ffffff">
                            INTEGRACION DE EXPEDIENTE UNICO DE CRÉDITO
                        </td>
                    </tr>
                </table>
            </div>
            
            <div class="header">{context['socio_info']}</div>

            <table>
                <thead>
                    <tr>
                        <th colspan="2" style="width: 50%; text-align: center; font-size: 15px; background-color:#d6dce4;">I. Identificación del Socio</th>
                        <th class="col-fecha" style="font-size: 15px; background-color:#d6dce4; padding: 5px 5px;">Fecha</th>
                        <th class="col-status" style="font-size: 15px; background-color:#d6dce4; padding: 5px 5px;">Status</th>
                        <th class="col-com" style="font-size: 15px; background-color:#d6dce4; padding: 5px 5px;">Comentarios</th>
                    </tr>
                </thead>
                <tbody>
        """

        for sec in secciones_data:
            html_string += f'<tr><td colspan="5" class="titulo-seccion">{sec["titulo"]}</td></tr>'
            for reg in sec['registros']:
                fecha = reg.fecha.strftime("%d/%m/%Y") if reg.fecha else ""
                html_string += f"""
                    <tr>
                        <td class="col-clave" style="text-align: center;">{reg.apartado.clave}</td>
                        <td class="col-desc">{reg.apartado.descripcion}</td>
                        <td class="col-fecha" style="text-align: center;">{fecha}</td>
                        <td class="col-status" style="text-align: center;">{reg.estatus or ''}</td>
                        <td class="col-com">{reg.comentario or ''}</td>
                    </tr>
                """

        html_string += "</tbody></table></body></html>"

        buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=buffer)

        if pisa_status.err:
            return HttpResponse("Error al renderizar PDF", status=500)

        buffer.seek(0)
        nombre_socio = "".join([c for c in str(expediente.socio.nombre) if c.isalnum() or c==' ']).replace(' ', '_')
        return FileResponse(buffer, as_attachment=True, filename=f"Expediente_{expediente.id}_{nombre_socio}.pdf")

    except Exception as e:
        return HttpResponse(f"Error critico al generar PDF: {str(e)}", status=500)



@login_required(login_url='/login/')    
def avances(request):
    mes_actual = datetime.now().month
    ano_actual = datetime.now().year
    
    mes_seleccionado = int(request.GET.get('mes', mes_actual))
    ano_seleccionado = int(request.GET.get('ano', ano_actual))
    tipo_reporte = request.GET.get('tipo_reporte', 'actual')

    if request.user.roles == 'Gerente Centro de Negocios':
        usuarios = User.objects.filter(roles__in=['Ejecutivo de Servicios','Gerente Centro de Negocios']).distinct()
        estados = Estado.objects.all().order_by('id')
    elif request.user.roles == 'Gerente de Credito':
        usuarios = User.objects.filter(roles__in=['Credito', 'Crédito', 'Gerente de Credito', 'Gerente de Crédito']).distinct()
        estados = Estado.objects.exclude(nombre='Nuevo').order_by('id')
    elif request.user.roles == 'Administrador':
        usuarios = User.objects.all()
        estados = Estado.objects.all().order_by('id')
    else:
        usuarios = User.objects.none()
        estados = Estado.objects.none()
    
    data_por_usuario = {}

    lista_usuarios = list(usuarios)
    lista_usuarios.append(None)

    for us in lista_usuarios:
        rol_limpio = request.user.roles
        es_credito = rol_limpio in ['Credito', 'Crédito', 'Gerente de Credito', 'Gerente de Crédito']

        if us is None:
            username_key = "Sin Asignar"
            grafica_barras = es_credito
            if es_credito:
                expedientes_totales = Expediente.objects.filter(usuarioCredito__isnull=True, eliminado=False)
            else:
                expedientes_totales = Expediente.objects.filter(usuario__isnull=True, eliminado=False)
        else:
            username_key = us.username
            rol_usuario = us.roles.strip() if us.roles else ""
            grafica_barras = rol_usuario in ['Credito', 'Crédito', 'Gerente de Credito', 'Gerente de Crédito']
            if grafica_barras:
                expedientes_totales = Expediente.objects.filter(usuarioCredito=us, eliminado=False)
            else:
                expedientes_totales = Expediente.objects.filter(usuario=us, eliminado=False)

        numExpedientes = expedientes_totales.count()
        
        conteo_por_estado = {}
        historico_cambios = {}
        estado_cierre_mes = {}
        
        for estado in estados:
            conteo_por_estado[estado.nombre] = {
                'count': 0,
                'color': estado.color,
                'id': estado.id
            }
            historico_cambios[estado.nombre] = 0
            estado_cierre_mes[estado.nombre] = 0

        if tipo_reporte == 'actual':
            for exp in expedientes_totales:
                if exp.estatus and exp.estatus.nombre in conteo_por_estado:
                    conteo_por_estado[exp.estatus.nombre]['count'] += 1
        else:
            expedientes_ids = expedientes_totales.values_list('id', flat=True)
            for exp_id in expedientes_ids:
                ultimo_cambio = EstadosFechas.objects.filter(
                    expediente_id=exp_id,
                    fecha__month=mes_seleccionado,
                    fecha__year=ano_seleccionado
                ).order_by('-fecha', '-hora').first()
                
                if ultimo_cambio:
                    if ultimo_cambio.estado.nombre in conteo_por_estado:
                        conteo_por_estado[ultimo_cambio.estado.nombre]['count'] += 1
                else:
                    ultimo_cambio_historico = EstadosFechas.objects.filter(
                        expediente_id=exp_id,
                        fecha__lt=datetime(ano_seleccionado, mes_seleccionado, 1).date()
                    ).order_by('-fecha', '-hora').first()
                    if ultimo_cambio_historico and ultimo_cambio_historico.estado.nombre in conteo_por_estado:
                        conteo_por_estado[ultimo_cambio_historico.estado.nombre]['count'] += 1

        historial_periodo = EstadosFechas.objects.filter(
            expediente__in=expedientes_totales,
            fecha__month=mes_seleccionado,
            fecha__year=ano_seleccionado
        )

        for registro in historial_periodo:
            if registro.estado.nombre in historico_cambios:
                historico_cambios[registro.estado.nombre] += 1

        expedientes_ids = expedientes_totales.values_list('id', flat=True)
        for exp_id in expedientes_ids:
            ultimo_cambio = EstadosFechas.objects.filter(
                expediente_id=exp_id,
                fecha__month=mes_seleccionado,
                fecha__year=ano_seleccionado
            ).order_by('-fecha', '-hora').first()
            
            if ultimo_cambio and ultimo_cambio.estado.nombre in estado_cierre_mes:
                estado_cierre_mes[ultimo_cambio.estado.nombre] += 1
        
        if tipo_reporte == 'actual':
            expedientes_completados = expedientes_totales.filter(estatus__id=2).count()
        else:
            expedientes_completados = 0
            for exp_id in expedientes_ids:
                uc = EstadosFechas.objects.filter(expediente_id=exp_id, fecha__month=mes_seleccionado, fecha__year=ano_seleccionado).order_by('-fecha', '-hora').first()
                if uc and uc.estado.id == 2:
                    expedientes_completados += 1

        if numExpedientes > 0:
            porcentaje_completado = (expedientes_completados / numExpedientes) * 100
        else:
            porcentaje_completado = 0
            
        if numExpedientes > 0 or us is not None:
            data_por_usuario[username_key] = {
                'total': numExpedientes,
                'porcentaje': round(porcentaje_completado, 2),
                'estados': conteo_por_estado,
                'grafica_barras': grafica_barras,
                'historico_cambios': historico_cambios,
                'estado_cierre_mes': estado_cierre_mes
            }
            
    context = {
        'data_por_usuario': data_por_usuario,
        'estados_list': estados,
        'mes_seleccionado': mes_seleccionado,
        'ano_seleccionado': ano_seleccionado,
        'tipo_reporte': tipo_reporte,
        'meses_rango': range(1, 13),
        'anos_rango': range(datetime.now().year - 3, datetime.now().year + 1)
    }
    
    return render(request, 'Index/avancesLayout.html', context)
@login_required(login_url='/login/')
#@user_passes_test(is_admin)
def administrador(request):

 
    form = UserAdminPassForm()
    usuarios = User.objects.all().order_by('username')
    context = {
        "usuarios": usuarios,
        "current_user_id": request.user.id,
        "form":form,
    }
    return render(request, "Index/administradorPage.html", context)

@login_required(login_url='/login/')
#@user_passes_test(is_admin)
def alta_usuario(request):
    if request.method == 'POST':
        form = UserAdminPassForm(request.POST) 
        if form.is_valid():
            user = form.save() 

            return redirect('Index:administrador')
        else:
            usuarios = User.objects.all().order_by('username')
            
            context = {
                "usuarios": usuarios,
                "current_user_id": request.user.id,
                "form": form,  
                "show_modal_error": True 
            }
            return render(request, 'Index/administradorPage.html', context)
    else:
        return redirect('Index:administrador')


@login_required(login_url='/login/')
#@user_passes_test(is_admin)
def editar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    form = UserAdminForm(instance=usuario)
    formpassowrd = UserPasswordForm()
    context = {
        'form': form,
        'usuario': usuario,
        'formpassowrd':formpassowrd
    }
    return render(request, "Index/editar_usuario.html", context)

@login_required(login_url='/login/')
#@user_passes_test(is_admin)
def editar_usuario_datos(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = UserAdminForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, f"Usuario {usuario.username} actualizado correctamente.")
            return redirect('Index:administrador')
        else:

            formpassowrd = UserPasswordForm()
            context = {
                'form': form,
                'usuario': usuario,
                'formpassowrd':formpassowrd
            }
            return render(request, "Index/editar_usuario.html", context)

@login_required(login_url='/login/')
#@user_passes_test(is_admin)
def editar_usuario_contrasena(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = UserPasswordForm(request.POST) 
        
        if form.is_valid():
            nueva_contrasena = form.cleaned_data['nueva_contrasena']
            
            usuario.set_password(nueva_contrasena)
            usuario.save()
            
            messages.success(request, f"Contraseña del usuario {usuario.username} actualizada correctamente.")
            return redirect('Index:administrador')
        else:
            form_datos = UserAdminForm(instance=usuario)
            context = {
                'form': form_datos,
                'usuario': usuario,
                'formpassowrd':form
            }
            return render(request, "Index/editar_usuario.html", context)

""" 


Todo este apartado es para lineas


"""
#Templeate donde se ven todas las lineas
@login_required(login_url='/login/')    
def lineasLayout(request):
    lineas = Linea.objects.all().order_by('-id')

    usuarios = User.objects.all()
    
    contexto = {
        'lineas': lineas,
        'usuarios': usuarios,
    }
    return render(request, 'Index/lineasLayout.html',contexto)

#Metodo para filtrar las lineas
def filtrar_lineas_ajax(request):
    socio_query = request.GET.get('socio', '').strip()
    page_number = request.GET.get('page', 1)

    lineas = Linea.objects.all()

    if socio_query:

        filtro = Q(expediente__socio__nombre__icontains=socio_query)
        
        if socio_query.isdigit():
            filtro |= Q(expediente__socio__id=int(socio_query))
        
        lineas = lineas.filter(filtro).distinct()
    lineas = lineas.order_by('-id')

    paginator = Paginator(lineas, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        'lineas': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator
    }

    return render(request, 'Index/tablasLineas.html', context)

def filtrar_obligados_ajax(request):
    socio_query = request.GET.get('socio', '').strip()
    page_number = request.GET.get('page', 1)

    lineas = Linea.objects.all()

    if socio_query:

        filtro = Q(expediente__socio__nombre__icontains=socio_query)
        
        if socio_query.isdigit():
            filtro |= Q(expediente__socio__id=int(socio_query))
        
        lineas = lineas.filter(filtro).distinct()
    lineas = lineas.order_by('-id')

    paginator = Paginator(lineas, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        'lineas': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator
    }

    return render(request, 'Index/tablasLineas.html', context)

def filtrar_representantes_ajax(request):
    socio_query = request.GET.get('socio', '').strip()
    page_number = request.GET.get('page', 1)

    lineas = Linea.objects.all()

    if socio_query:

        filtro = Q(expediente__socio__nombre__icontains=socio_query)
        
        if socio_query.isdigit():
            filtro |= Q(expediente__socio__id=int(socio_query))
        
        lineas = lineas.filter(filtro).distinct()
    lineas = lineas.order_by('-id')

    paginator = Paginator(lineas, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        'lineas': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator
    }

    return render(request, 'Index/tablasLineas.html', context)


#Metodo para editar una linea
@login_required(login_url='/login/')
def editarLinea(request, id):
    linea = get_object_or_404(Linea, pk=id)
    exito = False

    if request.method == "POST":
        form = LineaCrearForm(request.POST, instance=linea)
        if form.is_valid():
            form.save()
            messages.success(request, f"Línea de {linea.expediente.socio.nombre} actualizada.")
            exito = True
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    else:
        form = LineaCrearForm(instance=linea)
    
    context = {'form': form, 'exito': exito}
    return render(request, 'Index/editarLinea.html', context)


def cargaInicial(request):
    if request.method == 'POST':
        try:
            archivo = request.FILES['archivo_excel']
            
            contenido_texto = archivo.read().decode('utf-8').splitlines()
            lector_csv = csv.reader(contenido_texto)
            
            next(lector_csv, None)
            
            for index, row in enumerate(lector_csv, start=2):
                print(f"Fila {index} leída: {row}")
                
                if not row or not any(field.strip() for field in row):
                    print(f"Fila {index} está vacía, saltando...")
                    continue
                    
                print(f"Intentando guardar: Kepler={row[0]}, Nombre={row[1]}, Tipo={row[2]}")
                Socio.objects.create(
                    numeroKepler=row[0],
                    nombre=row[1],
                    tipoPersona=row[2]
                )
                print(f"Fila {index} guardada con éxito.")
                
        except Exception as e:
            print("Se produjo un error durante la carga:", str(e))
            
    return redirect('Index:administrador')




@login_required(login_url='/login/')

def juntasIndex(request):
    juntas = Cita.objects.all().prefetch_related('expedientes').order_by('-dia')
    cita_form = CitaForm()

    juntas_con_forms = []
    for junta in juntas:
        juntas_con_forms.append({
            'cita': junta,
            'form_editar': CitaForm(instance=junta)
        })

    context = {
        'juntas_con_forms': juntas_con_forms,
        'cita_form': cita_form,
    }
    return render(request, 'Index/juntasLayout.html', context)
@login_required(login_url='/login/')

def crearJunta(request):
    if request.method == 'POST':
        form = CitaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('Index:juntasIndex')
@login_required(login_url='/login/')

def editarJunta(request, cita):
    instancia_cita = get_object_or_404(Cita, pk=cita)
    if request.method == 'POST':
        form = CitaForm(request.POST, instance=instancia_cita)
        if form.is_valid():
            form.save()
            return redirect('Index:juntasIndex')
@login_required(login_url='/login/')

def eliminarJunta(request, cita):
    if request.method == 'POST':
        Cita.objects.filter(id=cita).delete()
        return redirect('Index:juntasIndex')
    
@login_required(login_url='/login/')

def enviarExpediente(request, expedienteID,reenviado):
    expediente = get_object_or_404(Expediente, pk=expedienteID)
    if reenviado == 'True':
        correoParaRevision(expediente,reenviado)
        getEstado = Estado.objects.get(nombre='Recepción')
        expediente.estatus = getEstado
        expediente.save()
    elif reenviado == 'False':
        correoParaRevision(expediente,reenviado)
        getEstado = Estado.objects.get(nombre='Completo')
        expediente.estatus = getEstado
        expediente.save()

    darAlta(expediente,getEstado.nombre,request.user)
    print("Alta dada")
    return redirect('Index:editarExpediente', expediente.id)


@login_required(login_url='/login/')
def cambiarUsuarioNegocios(request, id):
    expediente = get_object_or_404(Expediente, pk=id)

    if request.method == "POST":
        nuevo_usuarios_id = request.POST.get("usuario")

        if nuevo_usuarios_id:
            expediente.usuarioNegocios_id = nuevo_usuarios_id
            expediente.save()


    return redirect('Index:editarExpediente', id=expediente.id)

@login_required(login_url='/login/')

def recepcionExpediente(request, expedienteID,observaciones):
    expediente = get_object_or_404(Expediente, pk=expedienteID)
    
  
    if observaciones == "S":
        asunto = "Entrega de Expedientes"
        getEstado = Estado.objects.get(nombre='Recepción')
        expediente.estatus = getEstado
        expediente.usuarioCredito = request.user
        expediente.save()
        usuarioNombre = expediente.usuarioCredito.nombreCompleto 

        darAlta(expediente,getEstado.nombre,request.user)
        cuerpo_html = f"""
    <html>
        <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; padding: 20px; border-radius: 10px;">
                <h2 style="color: #2c3e50;">Entrega de Expedientes</h2>
                <p>Buenas tardes, {expediente.usuario.nombreCompleto}</p>
                <p>Confirmamos por este medio de haber recibido el expediente {expediente.id} del socio {expediente.socio.nombre} </p>
                <p> Quedando pendiente la revisión de los expedientes para su resguardo o el correo de faltantes. </p>
                <p> Quedo en espera de cualquier duda o comentario. </p>
 
                <p style="font-size: 0.9em; color: #555;">
                    Atentamente,<br>
                    <strong>{usuarioNombre}</strong>
                </p>
                <hr style="border: 0; border-top: 1px solid #eee;">
                <p style="font-size: 0.8em; color: #999; text-align: center;">
                    Este es un mensaje automático, por favor no responda a este correo.
                </p>
            </div>
        </body>
    </html>
    """

    elif observaciones == "N":
        getEstado = Estado.objects.get(nombre='Recepción con observaciones')
        expediente.estatus = getEstado
        expediente.usuarioCredito = request.user
        expediente.save()
        usuarioNombre = expediente.usuarioCredito.nombreCompleto 

        darAlta(expediente,getEstado.nombre,request.user)
        dominio = "http://192.168.0.29:8000/expedientes/editarExpediente/"
        url_final = f"{dominio}{expediente.id}/"
        asunto = f"RECHAZO: Expediente {expediente.id} - {expediente.socio.nombre}"
        seccionesConComentarios = []
        secciones = SeccionesExpediente.objects.filter(expediente=expediente).order_by('tipoDeSeccion', 'pk')

        for seccion in secciones:
            apartados = ApartadoCatalogo.objects.filter(tipoDeSeccion=seccion.tipoDeSeccion).order_by('clave')
            for apartado in apartados:
                registro = RegistroSeccion.objects.filter(seccion=seccion, apartado=apartado).first()

                if registro is not None:
                    if registro.comentarioCredito:
                        seccionesConComentarios.append({
                            'nombreSeccion': str(seccion.tituloSeccion),
                            'claveApartado': str(apartado.clave),
                            'comentario': str(registro.comentarioCredito)
                        })
                        registro.estatus = 'rechazado'
                        registro.save()

        
        comentarios_texto = ""
        for item in seccionesConComentarios:
            comentarios_texto += f"\n- [{item['nombreSeccion']} - {item['claveApartado']}]: {item['comentario']}"



        cuerpo_html = f"""
        <html>
            <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
                <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; padding: 20px; border-radius: 10px;">
                    <h2 style="color: #2c3e50;">Actualización de estatus del expediente</h2>
                    <p>Estimado usuario,</p>
                    <p>Le informamos que el expediente {expediente.id} del socio {expediente.socio.nombre} - {expediente.socio.numeroKepler} ha sido recepcionado con observaciones.</p>
                    <p>Atte {usuarioNombre}.</p>
                    <div style="margin: 30px 0; text-align: center;">
                        <a href="{url_final}" 
                           style="background-color: #007bff; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                            Revisar Expediente
                        </a>
                    </div>
                    <div style="background-color: #fdf2f2; border-left: 4px solid #f5c6cb; padding: 15px; margin: 20px 0; border-radius: 4px;">
                        <h4 style="margin-top: 0; color: #a94442;">Detalle de observaciones por sección:</h4>
                        <ul style="margin-bottom: 0; padding-left: 20px;">
        """

        for item in seccionesConComentarios:
            cuerpo_html += f"""
                            <li style="margin-bottom: 10px;">
                                <strong>{item['nombreSeccion']} (Clave: {item['claveApartado']}):</strong> {item['comentario']}
                            </li>
            """

        cuerpo_html += f"""
                        </ul>
                    </div>

                    <hr style="border: 0; border-top: 1px solid #eee;">
                    <p style="font-size: 0.8em; color: #999; text-align: center;">
                        Este es un mensaje automático, por favor no responda a este correo.
                    </p>
                </div>
            </body>
        </html>
        """


    else:
        print("Error")
        
 

    destinatario = [expediente.usuario.email] #,expediente.usuario.email 'daudiffred@ucg.com.mx'  expediente.usuarioCredito.email,
    
    
    
    SMTP_HOST = "ucg.com.mx"
    SMTP_PORT = 587
    USUARIO = "informacion@ucg.com.mx"
    CONTRASENA = "UcG911_@!#"

    mensaje = email.message.EmailMessage()
    mensaje["From"] = USUARIO
    mensaje["To"] = ", ".join(destinatario) if isinstance(destinatario, list) else destinatario
    mensaje["Subject"] = asunto
    
    mensaje.add_alternative(cuerpo_html, subtype="html")

    try:
        servidor = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        servidor.starttls()
        servidor.login(USUARIO, CONTRASENA)
        servidor.send_message(mensaje)
        servidor.quit()
        print("Correo enviado correctamente")
    except Exception as e:
        print("Error al enviar correo:", e)
    return redirect('Index:editarExpediente', expedienteID)

def enviarInvitacionCita(cita, expediente, usuario_accion=None, esCancelacion=False, esSugerencia=False):
    fechaInicio = datetime.combine(cita.dia, cita.hora)
    fechaFin = fechaInicio + timedelta(hours=1)

    cal = Calendar()
    cal.add('prodid', '-//Plataforma de Expedientes//mx//')
    cal.add('version', '2.0')
    
    metodo = 'CANCEL' if esCancelacion else 'REQUEST'
    cal.add('method', metodo)
    
    evento = Event()
    evento.add('dtstart', fechaInicio)
    evento.add('dtend', fechaFin)
    evento.add('dtstamp', datetime.now())
    evento.add('uid', f"cita-{cita.id}-{expediente.id}@plataformaexpedientes.com")
    
    rol_creador = getattr(usuario_accion, 'roles', None) if usuario_accion else None

    if esCancelacion:
        evento.add('status', 'CANCELLED')
        evento.add('summary', f"CANCELACIÓN: CITA PARA ENTREGAR EXPEDIENTE DEL SOCIO: {expediente.socio.numeroKepler} {expediente.socio.nombre.upper()}")
        evento.add('description', "Hola, se ha cancelado la cita que estaba programada para este expediente.")
        asunto = f"CANCELACIÓN: CITA PARA ENTREGAR EXPEDIENTE DEL SOCIO: {expediente.socio.numeroKepler} {expediente.socio.nombre.upper()} -- {cita.dia}"
    else:
        evento.add('status', 'CONFIRMED')
        evento.add('summary', f"CITA PARA ENTREGAR EXPEDIENTE DEL SOCIO: {expediente.socio.numeroKepler} {expediente.socio.nombre.upper()}")
        evento.add('description', "Cita programada para el expediente del socio.")
        
        if esSugerencia:
            asunto = f"Nueva Sugerencia de Cita: {cita.dia} a las {cita.hora} hrs"
            titulo_seccion = "Nueva Propuesta de Horario"
            texto_explicativo = f"Hola, se ha sugerido un cambio de horario para la entrega del expediente del socio <strong>{expediente.socio.numeroKepler} - {expediente.socio.nombre.upper()}</strong> debido a que el horario anterior fue rechazado."
        else:
            asunto = f"Confirmación de Cita: {cita.dia} a las {cita.hora} hrs"
            titulo_seccion = "Confirmación de Cita Programada"
            texto_explicativo = f"Hola, se ha agendado una cita para la entrega del expediente del socio <strong>{expediente.socio.numeroKepler} - {expediente.socio.nombre.upper()}</strong>."

    cal.add_component(evento)
    
    SMTP_HOST = "ucg.com.mx"
    SMTP_PORT = 465
    USUARIO = "informacion@ucg.com.mx"
    CONTRASENA = "UcG911_@!#"
    
    destinatariosCredito = ['portiz@ucg.com.mx', 'dcorrea@ucg.com.mx', 'mrubio@ucg.com.mx'] 
    destinatariosRaw = [
        getattr(getattr(expediente, 'usuarioCredito', None), 'email', None),
        getattr(getattr(expediente, 'usuario', None), 'email', None),
    ]

    destinatarios = list(set([correo for correo in destinatariosRaw + destinatariosCredito if correo]))
    
    # Lista de pruebas
    # destinatarios = ['daudiffred@ucg.com.mx']
    
    if not destinatarios:
        return

    try:
        conexion = get_connection(
            host=SMTP_HOST,
            port=SMTP_PORT,
            username=USUARIO,
            password=CONTRASENA,
            use_tls=False,
            use_ssl=True,
            timeout=10
        )
        
        urlConfirmar = "http://192.168.0.29:8000/expedientes/editarExpediente/Aceptar/"
        urlRechazar = "http://192.168.0.29:8000/expedientes/editarExpediente/Rechazar/"
        url_finalurlConfirmar = f"{urlConfirmar}{expediente.id}"
        url_finalurlurlRechazar = f"{urlRechazar}{expediente.id}"

        for correoDestino in destinatarios:
            if esCancelacion:
                cuerpo = f"""
                <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 5px;">
                    <h2 style="color: #dc3545;">Cancelación de Cita</h2>
                    <p style="color: #555555; font-size: 16px; line-height: 1.5;">
                        Hola, se ha cancelado la cita que estaba programada para la entrega del expediente del socio 
                        <strong>{expediente.socio.numeroKepler} - {expediente.socio.nombre.upper()}</strong>.
                    </p>
                    <div style="background-color: #f9f9f9; padding: 15px; border-radius: 5px; margin: 20px 0;">
                        <p style="margin: 5px 0; color: #333333;"><strong>Fecha Cancelada:</strong> {cita.dia}</p>
                        <p style="margin: 5px 0; color: #333333;"><strong>Hora Cancelada:</strong> {cita.hora} hrs</p>
                    </div>
                    <p style="color: #888888; font-size: 12px; margin-top: 30px; text-align: center; border-top: 1px solid #eeeeee; padding-top: 15px;">
                        Este es un correo automático generado por la Plataforma de Expedientes.
                    </p>
                </div>
                """
            else:
                mostrar_botones = True
                if rol_creador in ['Credito', 'Gerente de Credito', 'Gerente de Crédito']:
                    if correoDestino in destinatariosCredito or correoDestino == getattr(getattr(expediente, 'usuarioCredito', None), 'email', None):
                        mostrar_botones = False
                elif rol_creador in ['Ejecutivo de Servicios', 'Gerente Centro de Negocios']:
                    if correoDestino == getattr(getattr(expediente, 'usuario', None), 'email', None) or correoDestino not in destinatariosCredito:
                        if correoDestino != getattr(getattr(expediente, 'usuarioCredito', None), 'email', None):
                            mostrar_botones = False

                if mostrar_botones:
                    bloque_interaccion = f"""
                    <p style="color: #555555; font-size: 15px;">Por favor, responde a esta propuesta interactuando con los siguientes botones:</p>
                    <div style="margin-top: 25px; text-align: center;">
                        <a href="{url_finalurlConfirmar}" style="background-color: #28a745; color: white; padding: 12px 25px; text-decoration: none; font-weight: bold; border-radius: 4px; margin-right: 15px; display: inline-block;">Confirmar Cita</a>
                        <a href="{url_finalurlurlRechazar}" style="background-color: #dc3545; color: white; padding: 12px 25px; text-decoration: none; font-weight: bold; border-radius: 4px; display: inline-block;">Rechazar / Sugerir Cambio</a>
                    </div>
                    """
                else:
                    bloque_interaccion = f"""
                    <p style="color: #666666; font-size: 14px; font-style: italic; background-color: #f0f0f0; padding: 10px; border-radius: 4px; border-left: 4px solid #a0a0a0;">
                        Esperando respuesta de confirmación o cambio por parte de la contraparte del departamento correspondiente.
                    </p>
                    """

                cuerpo = f"""
                <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 5px;">
                    <h2 style="color: #333333;">{titulo_seccion}</h2>
                    <p style="color: #555555; font-size: 16px; line-height: 1.5;">
                        {texto_explicativo}
                    </p>
                    <div style="background-color: #f9f9f9; padding: 15px; border-radius: 5px; margin: 20px 0;">
                        <p style="margin: 5px 0; color: #333333;"><strong>Fecha Propuesta:</strong> {cita.dia}</p>
                        <p style="margin: 5px 0; color: #333333;"><strong>Hora Propuesta:</strong> {cita.hora} hrs</p>
                    </div>
                    {bloque_interaccion}
                    <p style="color: #888888; font-size: 12px; margin-top: 30px; text-align: center; border-top: 1px solid #eeeeee; padding-top: 15px;">
                        Este es un correo automático generado por la Plataforma de Expedientes.
                    </p>
                </div>
                """

            email = EmailMessage(
                subject=asunto,
                body=cuerpo,
                from_email=USUARIO,
                to=[correoDestino],
                connection=conexion
            )
            email.content_subtype = "html"
            
            ics_contenido = cal.to_ical().decode('utf-8')
            parte_calendario = SafeMIMEText(ics_contenido, 'calendar', 'utf-8')
            parte_calendario.set_param('method', metodo)
            parte_calendario.set_param('charset', 'UTF-8')
            
            email.attach(parte_calendario)
            email.attach(
                'invitacion-cita.ics',
                cal.to_ical(),
                f'text/calendar; charset=UTF-8; method={metodo}'
            )
            
            email.send()
        
        print("Correos enviados de forma individual correctamente")
        
    except Exception as e:
        print("Error al enviar correo:", e)
@login_required(login_url='/login/')
def asociarCitaExistente(request, expedienteId, citaId):
    if request.method == "POST":
        expediente = get_object_or_404(Expediente, pk=expedienteId)
        cita = get_object_or_404(Cita, pk=citaId)
        cita.expedientes.add(expediente)
        
        enviarInvitacionCita(cita, expediente, usuario_accion=request.user, esCancelacion=False, esSugerencia=False)
        messages.success(request, 'El expediente fue vinculado y se envió la invitación a Outlook.')
        
    return redirect('Index:editarExpediente', expedienteId)


@login_required(login_url='/login/')
def desasociarCitaExistente(request, expedienteId, citaId):
    if request.method == "POST":
        expediente = get_object_or_404(Expediente, pk=expedienteId)
        cita = get_object_or_404(Cita, pk=citaId)
        
        enviarInvitacionCita(cita, expediente, usuario_accion=request.user, esCancelacion=True, esSugerencia=False)
        
        cita.expedientes.remove(expediente)
        
        if not cita.expedientes.exists():
            cita.estatusCN = 'Pendiente'
            cita.estatusCR = 'Pendiente'
            cita.dia = None
            cita.hora = None
            cita.save()
            
        messages.success(request, 'La cita fue desasignada y se envió la notificación de cancelación.')
        
    return redirect('Index:editarExpediente', expedienteId)


@login_required(login_url='/login/')
def rechazarCita(request, id):
    expediente = get_object_or_404(Expediente, pk=id)
    cita = Cita.objects.filter(
        Q(expedientes=expediente),
    ).order_by('dia', 'hora').last()      
    
    if request.method == 'POST':
        form = CitaForm(request.POST, instance=cita, user=request.user)
        if form.is_valid():
            cita_modificada = form.save(commit=False)
            rol_usuario = getattr(request.user, 'roles', None)
            
            if rol_usuario in ['Ejecutivo de Servicios', 'Gerente Centro de Negocios']:
                cita_modificada.estatusCN = 'Aceptada'
                cita_modificada.estatusCR = 'Pendiente'
            elif rol_usuario in ['Credito', 'Gerente de Credito', 'Gerente de Crédito']:
                cita_modificada.estatusCR = 'Aceptada'
                cita_modificada.estatusCN = 'Pendiente'
            
            cita_modificada.save()
            form.save_m2m()
            
            enviarInvitacionCita(cita_modificada, expediente, usuario_accion=request.user, esCancelacion=False, esSugerencia=True)
            return redirect(reverse('Index:editarExpediente', args=[expediente.pk]))
    else:
        form = CitaForm(instance=cita, user=request.user)
        form.fields['expedientes'].widget = forms.MultipleHiddenInput()
        
    context = { 
        'expediente': expediente,
        'cita': cita,
        'form': form,
    }
    return render(request, 'Index/citaRechazada.html', context)


def crearCita(request, id):
    if request.method == "POST":
        cita_form = CitaForm(request.POST)
        if cita_form.is_valid():
            expediente = get_object_or_404(Expediente, pk=id)
            cita = cita_form.save(commit=False)
            cita.usuario = request.user
            
            rol_usuario = getattr(request.user, 'roles', None)
            if rol_usuario in ['Ejecutivo de Servicios', 'Gerente Centro de Negocios']:
                cita.estatusCN = 'Aceptada'
                cita.estatusCR = 'Pendiente'
            elif rol_usuario in ['Credito', 'Gerente de Credito', 'Gerente de Crédito']:
                cita.estatusCR = 'Aceptada'
                cita.estatusCN = 'Pendiente'
                
            cita.save()
            cita_form.save_m2m()
            cita.expedientes.add(expediente)
            
            enviarInvitacionCita(cita, expediente, usuario_accion=request.user, esCancelacion=False, esSugerencia=False)
            
    return redirect('Index:editarExpediente', id)
@login_required(login_url='/login/')

def aceptarCita(request, id):
    expediente = get_object_or_404(Expediente, pk=id)
    usuario = request.user
    cita = Cita.objects.filter(
        Q(expedientes=expediente),
    ).order_by('dia', 'hora').last()  
    
    if usuario.roles in ['Gerente de Credito', 'Credito']:
        cita.estatusCR = 'Aceptada'
    elif usuario.roles in ['Gerente Centro de Negocios', 'Ejecutivo de Servicios']:
        cita.estatusCN = 'Aceptada'
   
    cita.save()
    
    if cita.estatusCR == 'Aceptada' and cita.estatusCN == 'Aceptada':
        enviarCorreoCitaAceptada(cita, expediente)
        
    context = { 
        'expediente': expediente,
        'cita': cita,
    }
    return render(request, 'Index/citaAceptada.html', context)
def enviarCorreoCitaAceptada(cita, expediente):
    asunto = f"Cita Confirmada : {cita.dia} a las {cita.hora} hrs"
    
    cuerpo = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 5px;">
        <h2 style="color: #28a745;">Cita Confirmada</h2>
        <p style="color: #555555; font-size: 16px; line-height: 1.5;">
            Notificamos que ambos departamentos (Crédito y Centro de Negocios) están de acuerdo con el horario para la entrega y revisión del expediente del socio 
            <strong>{expediente.socio.numeroKepler} - {expediente.socio.nombre.upper()}</strong>.
        </p>
        <div style="background-color: #f1fcf4; padding: 15px; border: 1px solid #c3e6cb; border-radius: 5px; margin: 20px 0;">
            <p style="margin: 5px 0; color: #155724;"><strong>Fecha de la Cita:</strong> {cita.dia}</p>
            <p style="margin: 5px 0; color: #155724;"><strong>Hora de la Cita:</strong> {cita.hora} hrs</p>
            <p style="margin: 5px 0; color: #155724;"><strong>Estatus Crédito:</strong> Aceptada</p>
            <p style="margin: 5px 0; color: #155724;"><strong>Estatus Centro de Negocios:</strong> Aceptada</p>
        </div>
        <p style="color: #555555; font-size: 15px;">Este correo sirve como confirmación de que el horario ha quedado formalmente agendado en el sistema.</p>
        <p style="color: #888888; font-size: 12px; margin-top: 30px; text-align: center; border-top: 1px solid #eeeeee; padding-top: 15px;">
            Este es un correo automático generado por la Plataforma de Expedientes.
        </p>
    </div>
    """
    
    SMTP_HOST = "ucg.com.mx"
    SMTP_PORT = 465
    USUARIO = "informacion@ucg.com.mx"
    CONTRASENA = "UcG911_@!#"
    
    destinatariosCredito = ['portiz@ucg.com.mx','dcorrea@ucg.com.mx','mrubio@ucg.com.mx'] 
    destinatariosRaw = [
        getattr(getattr(expediente, 'usuarioCredito', None), 'email', None),
        getattr(getattr(expediente, 'usuario', None), 'email', None),
    ]

    destinatarios = list(set([correo for correo in destinatariosRaw + destinatariosCredito if correo]))
    #destinatarios = ['daudiffred@ucg.com.mx']
    try:
        conexion = get_connection(
            host=SMTP_HOST,
            port=SMTP_PORT,
            username=USUARIO,
            password=CONTRASENA,
            use_tls=False,
            use_ssl=True,
            timeout=10
        )
        
        for correoDestino in destinatarios:
            email = EmailMessage(
                subject=asunto,
                body=cuerpo,
                from_email=USUARIO,
                to=[correoDestino],
                connection=conexion
            )
            email.content_subtype = "html"
            email.send()
        
        print("Correo informativo de confirmación enviado con éxito")
        
    except Exception as e:
        print("Error al enviar correo de aceptación:", e)
         
@login_required(login_url='/login/')

def agregarRenglonExpediente(request, expedienteID, seccionID, apartadoID):
    expediente = get_object_or_404(Expediente, pk=expedienteID)
    seccion_actual = get_object_or_404(SeccionesExpediente, pk=seccionID)
    apartado_origen = get_object_or_404(ApartadoCatalogo, pk=apartadoID)
    
    registros_existentes = RegistroSeccion.objects.filter(
        seccion=seccion_actual, 
        apartado=apartado_origen
    )
    
    registro_origen = registros_existentes.first()
    
    if not registro_origen:
        return redirect('Index:editarExpediente', expedienteID)
        
    max_secuencial = registros_existentes.aggregate(models.Max('secuencial'))['secuencial__max'] or 0
    nuevo_secuencial = max_secuencial + 1
    
    nuevo_registro = RegistroSeccion.objects.create(
        seccion=seccion_actual,
        apartado=apartado_origen,
        estatus=registro_origen.estatus,
        comentario=registro_origen.comentario,
        comentarioCredito=registro_origen.comentarioCredito,
        es_fecha=registro_origen.es_fecha,
        secuencial=nuevo_secuencial
    )
    
    return redirect('Index:editarExpediente', expedienteID)
@login_required(login_url='/login/')

def eliminarRenglonExpediente(request, expedienteID, seccionID, apartadoID):
    expediente = get_object_or_404(Expediente, pk=expedienteID)
    seccion_actual = get_object_or_404(SeccionesExpediente, pk=seccionID)
    apartado_origen = get_object_or_404(ApartadoCatalogo, pk=apartadoID)
    
    ultimo_registro = RegistroSeccion.objects.filter(
        seccion=seccion_actual,
        apartado=apartado_origen
    ).order_by('-secuencial').first()
    
    if ultimo_registro and ultimo_registro.secuencial > 1:
        ultimo_registro.delete()
        messages.success(request, 'Renglón extra eliminado con éxito.')
    else:
        messages.error(request, 'No es posible eliminar el renglón principal del apartado.')
        
    return redirect('Index:editarExpediente', expedienteID)


@login_required(login_url='/login/')
def revisionExpediente(request, expedienteID,observaciones):
    expediente = get_object_or_404(Expediente, pk=expedienteID)
    
  
    if observaciones == "S":
        getEstado = Estado.objects.get(nombre='En revisión')
        expediente.estatus = getEstado
        expediente.save()

        darAlta(expediente,getEstado.nombre,request.user)

    elif observaciones == "N":
        print("aqui")
        getEstado = Estado.objects.get(nombre='En revisión con observaciones')
        expediente.estatus = getEstado
        expediente.save()
        darAlta(expediente,getEstado.nombre,request.user)

    elif observaciones == "R":
        from email.message import EmailMessage

        getEstado = Estado.objects.get(nombre='Recepción con observaciones')
        expediente.estatus = getEstado
        expediente.save()

        darAlta(expediente,getEstado.nombre,request.user)
        seccionesConComentarios = []
        secciones = SeccionesExpediente.objects.filter(expediente=expediente).order_by('tipoDeSeccion', 'pk')

        for seccion in secciones:
            apartados = ApartadoCatalogo.objects.filter(tipoDeSeccion=seccion.tipoDeSeccion).order_by('clave')
            for apartado in apartados:
                registro = RegistroSeccion.objects.filter(seccion=seccion, apartado=apartado).first()

                if registro is not None:
                    if registro.comentarioCredito:
                        seccionesConComentarios.append({
                            'nombreSeccion': str(seccion.tituloSeccion),
                            'claveApartado': str(apartado.clave),
                            'comentario': str(registro.comentarioCredito)
                        })

        destinatario = [expediente.usuario.email] #expediente.usuarioCredito.email
        #destinatario = ["daudiffred@ucg.com.mx"]    
        dominio = "http://192.168.0.29:8000/expedientes/editarExpediente/"
        url_final = f"{dominio}{expediente.id}/"
        asunto = f"RECHAZO EN RECEPCIÓN: Expediente {expediente.id} - {expediente.socio.nombre}"

        comentarios_texto = ""
        for item in seccionesConComentarios:
            comentarios_texto += f"\n- [{item['nombreSeccion']} - {item['claveApartado']}]: {item['comentario']}"


        cuerpo_html = f"""
        <html>
            <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
                <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; padding: 20px; border-radius: 10px;">
                    <h2 style="color: #2c3e50;">Actualización de estatus del expediente</h2>
                    <p>Estimado usuario,</p>
                    <p>Le informamos que el expediente {expediente.id} del socio {expediente.socio.nombre} - {expediente.socio.numeroKepler} ha sido rechazado durante su recepción.</p>
                    <p>Atte {expediente.usuarioCredito.username}.</p>
                    <div style="margin: 30px 0; text-align: center;">
                        <a href="{url_final}" 
                           style="background-color: #007bff; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                            Revisar Expediente
                        </a>
                    </div>
                    <div style="background-color: #fdf2f2; border-left: 4px solid #f5c6cb; padding: 15px; margin: 20px 0; border-radius: 4px;">
                        <h4 style="margin-top: 0; color: #a94442;">Detalle de observaciones por sección:</h4>
                        <ul style="margin-bottom: 0; padding-left: 20px;">
        """

        for item in seccionesConComentarios:
            cuerpo_html += f"""
                            <li style="margin-bottom: 10px;">
                                <strong>{item['nombreSeccion']} (Clave: {item['claveApartado']}):</strong> {item['comentario']}
                            </li>
            """

        cuerpo_html += f"""
                        </ul>
                    </div>

                    <hr style="border: 0; border-top: 1px solid #eee;">
                    <p style="font-size: 0.8em; color: #999; text-align: center;">
                        Este es un mensaje automático, por favor no responda a este correo.
                    </p>
                </div>
            </body>
        </html>
        """
        SMTP_HOST = "ucg.com.mx"
        SMTP_PORT = 587
        USUARIO = "informacion@ucg.com.mx"
        CONTRASENA = "UcG911_@!#"

        mensaje = EmailMessage()
        mensaje["From"] = USUARIO
        mensaje["To"] = ", ".join(destinatario) if isinstance(destinatario, list) else destinatario
        mensaje["Subject"] = asunto

        mensaje.add_alternative(cuerpo_html, subtype="html")

        try:
            servidor = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
            servidor.starttls()
            servidor.login(USUARIO, CONTRASENA)
            servidor.send_message(mensaje)
            servidor.quit()
            print("Correo enviado correctamente")
        except Exception as e:
            print("Error al enviar correo:", e)
    else:
            print("Error")
        
        
    return redirect('Index:editarExpediente', expedienteID)
@login_required(login_url='/login/')
def lineaEliminar(request, expediente_id, linea_id):
    if request.method == "POST":
        linea = get_object_or_404(Linea, pk=linea_id, expediente_id=expediente_id)
        
        try:
            with transaction.atomic():
                secciones_linea = SeccionesExpediente.objects.filter(linea=linea, expediente_id=expediente_id)
                RegistroSeccion.objects.filter(seccion__in=secciones_linea).delete()
                secciones_linea.delete()
                linea.delete()
                
            messages.success(request, 'Línea de crédito y todos sus apartados asociados fueron eliminados con éxito.')
        except Exception as e:
            messages.error(request, f'Error al eliminar la línea: {str(e)}')
            
        return redirect(reverse('Index:editarExpediente', args=[expediente_id]))
        
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)



def eliminarRepresentante(request,rep,exp):
    try:
        expediente = Expediente.objects.get(id=exp)
        obligado = RepresentanteLegal.objects.get(id=rep)
        
        try:
            seccion_obl = SeccionesExpediente.objects.get(
                expediente=expediente,
                tipoDeSeccion='B',
                tituloSeccion=f"Representante legal - {obligado.nombre}"
            )
            seccion_obl.delete()
        except SeccionesExpediente.DoesNotExist:
            pass

        obligado.expedientes.remove(expediente)

    except (Expediente.DoesNotExist, RepresentanteLegal.DoesNotExist):
        raise Http404

    return redirect(reverse('Index:editarExpediente', args=[expediente.id]))




def eliminarObligado(request, obl, exp):
    try:
        expediente = Expediente.objects.get(id=exp)
        obligado = ObligadoSolidario.objects.get(id=obl)
        
        try:
            seccion_obl = SeccionesExpediente.objects.get(
                expediente=expediente,
                tipoDeSeccion='C',
                tituloSeccion=f"Obligado solidario y garantes - {obligado.nombre}"
            )
            seccion_obl.delete()
        except SeccionesExpediente.DoesNotExist:
            pass

        obligado.expedientes.remove(expediente)

    except (Expediente.DoesNotExist, ObligadoSolidario.DoesNotExist):
        raise Http404

    return redirect(reverse('Index:editarExpediente', args=[expediente.id]))

def checkBoxChange(request, seccion, apartado, secuencial):
    registro = get_object_or_404(
        RegistroSeccion, 
        seccion_id=seccion, 
        apartado_id=apartado, 
        secuencial=secuencial
    )
    registro.enviar = not registro.enviar
    registro.save()
    return JsonResponse({'success': True, 'enviar': registro.enviar})

def notificarFaltantes(request,expedienteID):
    from email.message import EmailMessage
    expediente = get_object_or_404(Expediente, pk=expedienteID)

    seccionesConCheckBox = []
    secciones = SeccionesExpediente.objects.filter(expediente=expediente).order_by('tipoDeSeccion', 'pk')
    
    for seccion in secciones:
        apartados = ApartadoCatalogo.objects.filter(tipoDeSeccion=seccion.tipoDeSeccion).order_by('clave')
        for apartado in apartados:
            registro = RegistroSeccion.objects.filter(seccion=seccion, apartado=apartado).first()
            
            if registro is not None:
                if registro.enviar:
                    seccionesConCheckBox.append({
                        'nombreSeccion': str(seccion.tituloSeccion),
                        'claveApartado': str(apartado.clave),
                    })
                    
    destinatario = [expediente.usuario.email,expediente.usuarioNegocios.email]     
    dominio = "http://192.168.0.29:8000/expedientes/editarExpediente/"
    url_final = f"{dominio}{expediente.id}/"
    asunto = f"Documentos faltantes: Expediente {expediente.id} - {expediente.socio.nombre}"
    
    comentarios_texto = ""
    for item in seccionesConCheckBox:
        comentarios_texto += f"\n- [{item['nombreSeccion']} - {item['claveApartado']}]: "


    cuerpo_html = f"""
    <html>
        <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; padding: 20px; border-radius: 10px;">
                <h2 style="color: #2c3e50;">Actualización de estatus del expediente</h2>
                <p>Estimado usuario,</p>
                <p>Le informamos que el expediente {expediente.id} del socio {expediente.socio.nombre} - {expediente.socio.numeroKepler} requiere ciertos documentos.</p>
                <p>Atte {expediente.usuario.username}.</p>
                <div style="margin: 30px 0; text-align: center;">
                    <a href="{url_final}" 
                       style="background-color: #007bff; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                        Revisar Expediente
                    </a>
                </div>
                <div style="background-color: #fdf2f2; border-left: 4px solid #f5c6cb; padding: 15px; margin: 20px 0; border-radius: 4px;">
                    <h4 style="margin-top: 0; color: #a94442;">Detalle de observaciones por sección:</h4>
                    <ul style="margin-bottom: 0; padding-left: 20px;">
    """

    for item in seccionesConCheckBox:
        cuerpo_html += f"""
                        <li style="margin-bottom: 10px;">
                            <strong>{item['nombreSeccion']} (Clave: {item['claveApartado']}):</strong> 
                        </li>
        """

    cuerpo_html += f"""
                    </ul>
                </div>
          
                <hr style="border: 0; border-top: 1px solid #eee;">
                <p style="font-size: 0.8em; color: #999; text-align: center;">
                    Este es un mensaje automático, por favor no responda a este correo.
                </p>
            </div>
        </body>
    </html>
    """
    SMTP_HOST = "ucg.com.mx"
    SMTP_PORT = 587
    USUARIO = "informacion@ucg.com.mx"
    CONTRASENA = "UcG911_@!#"
    
    mensaje = EmailMessage()
    mensaje["From"] = USUARIO
    mensaje["To"] = ", ".join(destinatario) if isinstance(destinatario, list) else destinatario
    mensaje["Subject"] = asunto
    
    mensaje.add_alternative(cuerpo_html, subtype="html")

    try:
        servidor = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        servidor.starttls()
        servidor.login(USUARIO, CONTRASENA)
        servidor.send_message(mensaje)
        servidor.quit()
        print("Correo enviado correctamente")
    except Exception as e:
        print("Error al enviar correo:", e)
    return redirect('Index:editarExpediente', expediente.id)



@login_required(login_url='/login/')
def procesarArchivos(request, id):
    if request.method != "POST":
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    expediente = get_object_or_404(Expediente, pk=id)
    secciones = SeccionesExpediente.objects.filter(expediente=expediente)
    
    kepler = str(expediente.socio.numeroKepler).strip() if expediente.socio.numeroKepler else "0000"
    nombre_socio = str(expediente.socio.nombre).strip().upper()
    rutaServidor = fr"\\192.168.0.96\intranetucg$$\Evidencias\652 Digitalización de expedientes de crédito"

    carpeta_socio = None
    try:
        if os.path.exists(rutaServidor):
            for nombre_dir in os.listdir(rutaServidor):
                if kepler != "0000" and kepler in nombre_dir:
                    carpeta_socio = os.path.join(rutaServidor, nombre_dir)
                    break
                elif kepler == "0000" and nombre_socio in nombre_dir.upper():
                    carpeta_socio = os.path.join(rutaServidor, nombre_dir)
                    break
        
        if not carpeta_socio:
            nombre_nueva_carpeta = f"{kepler} {nombre_socio}".strip()
            carpeta_socio = os.path.join(rutaServidor, nombre_nueva_carpeta)
            os.makedirs(carpeta_socio, exist_ok=True)
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error de acceso o creación de carpeta raíz: {str(e)}'}, status=500)

    rutaMaestra = os.path.join(carpeta_socio, "Maestra")
    rutaOperativa = os.path.join(carpeta_socio, "Operativa")

    mapeo_secciones = {
        "1": os.path.join(rutaMaestra, "I. Identificación del Socio"),
        "2": os.path.join(rutaMaestra, "II. Información Financiera"),
        "3": os.path.join(rutaOperativa, "III. Estudio de Crédito"),
        "4": os.path.join(rutaOperativa, "IV. Información de garantias"),
        "5": os.path.join(rutaOperativa, "V. Contratos"),
        "6": os.path.join(rutaOperativa, "VI. Seguimiento"),
        "7": os.path.join(rutaOperativa, "VII. Correspondencia")
    }
    
    meses_map = {
        "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
        "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12
    }
    
    meses_inv_map = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
    }

    archivos = request.FILES.getlist('archivos_pdf')
    resultados_proceso = []

    registros_validos = RegistroSeccion.objects.filter(seccion__in=secciones).select_related('apartado')
    mapeo_claves_registro = {}
    for r in registros_validos:
        clave_limpia = str(r.apartado.clave).strip()
        mapeo_claves_registro[clave_limpia] = r

    for f in archivos:
        nombre_archivo = f.name
        nombre_sin_ext, ext = os.path.splitext(nombre_archivo)
        
        partes_nombre = nombre_sin_ext.strip().split(' ')
        if not partes_nombre or partes_nombre[0] == "":
            resultados_proceso.append({'archivo': nombre_archivo, 'success': False, 'mensaje': 'Nombre inválido'})
            continue
            
        linea_detectada_num = None
        clave_detectada = None

        if '.' not in partes_nombre[0]:
            linea_detectada_num = partes_nombre[0].strip()
            if len(partes_nombre) > 1 and '.' in partes_nombre[1]:
                clave_detectada = partes_nombre[1].strip().rstrip('.')
            else:
                resultados_proceso.append({'archivo': nombre_archivo, 'success': False, 'mensaje': 'Estructura de clave x.xx no encontrada tras número de línea'})
                continue
        else:
            clave_detectada = partes_nombre[0].strip().rstrip('.')

        prefijo = clave_detectada.split('.')[0]
        
        ruta_destino_base = mapeo_secciones.get(prefijo)
        if not ruta_destino_base:
            resultados_proceso.append({'archivo': nombre_archivo, 'success': False, 'mensaje': f'Clave "{clave_detectada}" no mapea a Maestra/Operativa'})
            continue

        if prefijo in ["3", "4", "5", "6", "7"]:
            if linea_detectada_num:
                lineas_asociadas = Linea.objects.filter(expediente=expediente, numero=linea_detectada_num)
            else:
                lineas_asociadas = Linea.objects.filter(expediente=expediente)

            if lineas_asociadas.exists():
                os.makedirs(rutaOperativa, exist_ok=True)
                try:
                    subdirs_operativa = os.listdir(rutaOperativa)
                    encontrado = False
                    
                    for linea in lineas_asociadas:
                        prefijo_linea = f"{linea.numero} "
                        for subdir in subdirs_operativa:
                            if subdir.strip().startswith(prefijo_linea):
                                ruta_subdir_completa = os.path.join(rutaOperativa, subdir)
                                if os.path.isdir(ruta_subdir_completa):
                                    nombre_carpeta_seccion = os.path.basename(ruta_destino_base)
                                    ruta_destino_base = os.path.join(ruta_subdir_completa, nombre_carpeta_seccion)
                                    encontrado = True
                                    break
                        if encontrado:
                            break
                    
                    if not encontrado:
                        linea_primera = lineas_asociadas.first()
                        num_l = str(linea_primera.numero).strip()
                        kepler_l = kepler
                        abrev_l = str(linea_primera.abreviacion).strip() if linea_primera.abreviacion else ""
                        monto_l = f"{linea_primera.monto:,}"
                        
                        fecha_str = ""
                        if linea_primera.fecha:
                            m_str = meses_inv_map.get(linea_primera.fecha.month, "")
                            y_str = linea_primera.fecha.strftime("%Y")
                            fecha_str = f"{m_str} {y_str}".strip()
                        
                        nombre_nueva_linea_dir = f"{num_l} {kepler_l} {abrev_l} ${monto_l} {fecha_str}".strip()
                        ruta_nueva_linea_completa = os.path.join(rutaOperativa, nombre_nueva_linea_dir)
                        os.makedirs(ruta_nueva_linea_completa, exist_ok=True)
                        
                        nombre_carpeta_seccion = os.path.basename(ruta_destino_base)
                        ruta_destino_base = os.path.join(ruta_nueva_linea_completa, nombre_carpeta_seccion)
                        
                except Exception:
                    pass

        registro_asociado = mapeo_claves_registro.get(clave_detectada)
        if not registro_asociado:
            partes_clave = clave_detectada.split('.')
            if len(partes_clave) == 2 and len(partes_clave[1]) > 2:
                clave_truncada = f"{partes_clave[0]}.{partes_clave[1][:2]}"
                registro_asociado = mapeo_claves_registro.get(clave_truncada)

        if not registro_asociado:
            resultados_proceso.append({'archivo': nombre_archivo, 'success': False, 'mensaje': f'No hay renglón en base de datos para la clave {clave_detectada}'})
            continue

        match_fecha = re.search(r'([a-z]{3})\s+(\d{2})$', nombre_sin_ext.strip().lower())
        fecha_procesada = None
        
        if match_fecha:
            mes_texto = match_fecha.group(1)
            anio_texto = match_fecha.group(2)
            if mes_texto in meses_map:
                try:
                    fecha_procesada = datetime(int(f"20{anio_texto}"), meses_map[mes_texto], 1).date()
                except ValueError:
                    pass

        if not fecha_procesada:
            resultados_proceso.append({'archivo': nombre_archivo, 'success': False, 'mensaje': 'Estructura de fecha incorrecta al final del nombre'})
            continue

        try:
            os.makedirs(ruta_destino_base, exist_ok=True)
            
            carpeta_destino_final = None
            for item in os.listdir(ruta_destino_base):
                ruta_item = os.path.join(ruta_destino_base, item)
                if os.path.isdir(ruta_item):
                    item_clean = item.strip()
                    if item_clean.startswith(clave_detectada + " ") or item_clean.startswith(clave_detectada + ".") or item_clean == clave_detectada:
                        carpeta_destino_final = ruta_item
                        break
            
            if not carpeta_destino_final:
                carpeta_destino_final = os.path.join(ruta_destino_base, clave_detectada)
                os.makedirs(carpeta_destino_final, exist_ok=True)
            
            nombre_sin_ext_limpio = nombre_sin_ext
            if linea_detectada_num and nombre_sin_ext_limpio.startswith(linea_detectada_num + " "):
                nombre_sin_ext_limpio = nombre_sin_ext_limpio[len(linea_detectada_num):].strip()

            nombre_final_archivo = f"{nombre_sin_ext_limpio}{ext}"
            if prefijo in ["3", "4", "5", "6", "7"] and registro_asociado.numero:
                str_numero = str(registro_asociado.numero).strip()
                if str_numero in nombre_sin_ext_limpio:
                    nombre_sin_ext_limpio = nombre_sin_ext_limpio.replace(str_numero, "").replace("  ", " ").strip()
                    nombre_final_archivo = f"{nombre_sin_ext_limpio}{ext}"

            ruta_completa_archivo = os.path.join(carpeta_destino_final, nombre_final_archivo)
            
            with open(ruta_completa_archivo, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk)
            
            registro_asociado.es_fecha = True
            registro_asociado.fecha = fecha_procesada
            registro_asociado.numero = None
            if not registro_asociado.estatus or registro_asociado.estatus == "":
                registro_asociado.estatus = "Completo"
            registro_asociado.save()
                    
            resultados_proceso.append({'archivo': nombre_final_archivo, 'success': True, 'mensaje': f'Copiado y enlazado con fecha {fecha_procesada.strftime("%b %y")}'})
        except Exception as e:
            resultados_proceso.append({'archivo': nombre_archivo, 'success': False, 'mensaje': f'Error en escritura: {str(e)}'})

    return JsonResponse({
        'global_success': any(r['success'] for r in resultados_proceso),
        'resultados': resultados_proceso
    })


@login_required(login_url='/login/')
def avancesMovimientos(request):
    expediente = Expediente.objects.all().order_by('socio__nombre')    
    todosEstados = EstadosFechas.objects.all().order_by('fecha')
    
    for exp in expediente:
        estados_exp = [e for e in todosEstados if e.expediente_id == exp.id]
        for i in range(len(estados_exp)):
            if i < len(estados_exp) - 1:
                fecha_actual = estados_exp[i].fecha
                fecha_siguiente = estados_exp[i+1].fecha
                diferencia = (fecha_siguiente - fecha_actual).days
                estados_exp[i].dias_siguiente = diferencia
            else:
                estados_exp[i].dias_siguiente = None

    contest = {
        'expediente': expediente,
        'todosEstados': todosEstados,
    }

    return render(request, 'Index/avancesMovimientos.html', contest)













